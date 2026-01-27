# ============================================================
# SISTEMA DE INTERNAÇÕES — VERSÃO SUPABASE (Cloud)
# Visual e fluxo do app "Versão Final" — DB: Supabase
# Melhorias: importação turbo, cache TTL centralizado, view opcional
# Ajuste: normalização de 'atendimento' (0007064233 == 7064233)
# + NOVO: Importação por (atendimento, aviso) com Regra B (filhas)
# ============================================================

import streamlit as st
import pandas as pd
from datetime import date, datetime
import io
import json
import re
import math, zipfile, io as _io, time
from typing import List, Dict, Any
from io import BytesIO
from collections import OrderedDict  # <<< NOVO
from collections import defaultdict

import streamlit.components.v1 as components

# ==== Supabase ====
from supabase import create_client, Client
from postgrest import APIError

# ==== PDF (ReportLab) - opcional ====
REPORTLAB_OK = True
try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
except ModuleNotFoundError:
    REPORTLAB_OK = False


# logo antes de "registros = parse_tiss_original(csv_text)"
import importlib
import parser as _parser_mod
importlib.reload(_parser_mod)
parse_tiss_original = _parser_mod.parse_tiss_original

# Parser (seu módulo)
# -> mantenha o arquivo parser.py no projeto com parse_tiss_original(csv_text) definido.
try:
    from parser import parse_tiss_original
except Exception:
    parse_tiss_original = None

# ============================================================
# SUPABASE — Conexão
# ============================================================
URL = st.secrets.get("SUPABASE_URL", "")
KEY = st.secrets.get("SUPABASE_KEY", "")
if not URL or not KEY:
    st.error("Configure SUPABASE_URL e SUPABASE_KEY em Secrets para iniciar o app.")
    st.stop()
supabase: Client = create_client(URL, KEY)

def _sb_debug_error(e: APIError, prefix="Erro Supabase"):
    st.error(prefix)
    with st.expander("Detalhes técnicos"):
        st.code(
            f"code: {getattr(e, 'code', None)}\n"
            f"message: {getattr(e, 'message', None)}\n"
            f"details: {getattr(e, 'details', None)}\n"
            f"hint: {getattr(e, 'hint', None)}",
            language="text",
        )

# ============================================================
# Configurações de Desempenho
# ============================================================
# TTLs centralizados (invalidados manualmente após CRUD)
TTL_LONG = 300  # 5 min (listas estáveis: hospitais)
TTL_MED = 180   # 3 min (bases agregadas das telas)
TTL_SHORT = 120 # 2 min (consultas frequentes)

def _to_bool(x):
    if isinstance(x, bool):
        return x
    s = str(x).strip().lower()
    return s in ("1", "true", "yes", "y", "on")

USE_DB_VIEW = _to_bool(st.secrets.get("USE_DB_VIEW", False))  # opcional: usar VIEW vw_procedimentos_internacoes

def invalidate_caches():
    """Invalida TODOS os caches (chamado após qualquer CRUD)."""
    try:
        st.cache_data.clear()
    except Exception:
        pass

# ============================================================
# Domínio e Aparência
# ============================================================
STATUS_OPCOES = [
    "Pendente",
    "Não Cobrar",
    "Enviado para pagamento",
    "Aguardando Digitação - AMHP",
    "Finalizado",
]
PROCEDIMENTO_OPCOES = ["Cirurgia / Procedimento", "Parecer"]
GRAU_PARTICIPACAO_OPCOES = ["Cirurgião", "1 Auxiliar", "2 Auxiliar", "3 Auxiliar", "Clínico"]
ALWAYS_SELECTED_PROS = {"JOSE.ADORNO", "CASSIO CESAR", "FERNANDO AND", "SIMAO.MATOS"}



def find_allowed_in_row(cells: list[str]) -> str | None:
    norm_cells = [str(c or "").strip().upper() for c in cells]
    for name in ALWAYS_SELECTED_PROS:
        if name.upper().strip() in norm_cells:
            return name
    return None

def choose_professional_for_group(group_rows: list[list[str]]):
    # group_rows: lista de linhas (mestre primeiro) -> cada linha é a lista de células do CSV
    # 1) Regra A (mestre)
    name = find_allowed_in_row(group_rows[0])
    if name:
        return name, "A"

    # 2) Regra B (primeira filha que contenha um dos nomes)
    for row in group_rows[1:]:
        name = find_allowed_in_row(row)
        if name:
            return name, "B"

    # 3) (opcional) fallback para o campo 'prestador' já parseado, se houver
    return "", "SKIP"


def inject_css():
    st.markdown("""
<style>
/* ===== KPIs maiores e centralizados ===== */
.kpi-wrap.center .kpi{ text-align:center; }
.kpi.big .label{ font-size: 1.05rem; font-weight: 700; }
.kpi.big .value{ font-size: 2.4rem; line-height: 2.6rem; font-weight: 800; color: var(--text); }
.kpi.big .hint{ font-size: .95rem; color: var(--muted); margin-top: 4px;}
.kpi.big:hover{ box-shadow: 0 1px 0 rgba(0,0,0,.03); }
.kpi-action button{ font-size: 0.95rem !important; font-weight: 700 !important; }
:root{ --bg-main:#F5F6F7; --bg-card:#FFFFFF; --border:#D0D7DE; --text:#24292F; --muted:#6B7280; --primary:#1F6FEB; --primary-hover:#1558B0; --radius:8px; }
html, body, .stApp{ background-color:var(--bg-main)!important; color:var(--text)!important; font-family:"Segoe UI", Roboto, Arial, sans-serif;}
.app-header{ background:var(--bg-main); padding:10px 12px; margin:-1.2rem -1rem .8rem -1rem; border-bottom:1px solid var(--border); }
.app-header .title{ font-size:1.2rem; font-weight:700; color:var(--primary); }
.app-header .sub{ font-size:.9rem; color:var(--muted); }
.soft-card{ background:var(--bg-card); border:1px solid var(--border); border-radius:var(--radius); padding:14px 16px; margin-bottom:12px; }
.stTextInput input,.stNumberInput input,.stDateInput input,.stTextArea textarea{ background:#FFF!important;color:var(--text)!important;border:1px solid var(--border)!important;border-radius:var(--radius)!important; box-shadow:none!important;}
label, .st-emotion-cache-1qg05tj p, .stMarkdown p{ color:var(--muted)!important; }
div[data-baseweb="select"]{ background:#FFF!important;border:1px solid var(--border)!important;border-radius:var(--radius)!important;color:var(--text)!important;}
div[data-baseweb="select"] div[role="combobox"]{ background:#FFF!important;color:var(--text)!important;}
div[data-baseweb="menu"]{ background:#FFF!important;border:1px solid var(--border)!important;border-radius:var(--radius)!important;color:var(--text)!important;}
div[data-baseweb="option"]{ background:#FFF!important;color:var(--text)!important;}
div[data-baseweb="option"][aria-selected="true"]{ background:#EEF2FF!important;color:#111827!important;}
div[data-baseweb="option"]:hover{ background:#F3F4F6!important;}
.stFileUploader > section{ border:1px solid var(--border)!important; background:#FFF!important;border-radius:var(--radius)!important;}
.stFileUploader div[role="button"]{ background:#FFF!important;color:var(--text)!important;border:1px solid var(--border)!important;border-radius:var(--radius)!important;}
.stButton>button{ background:var(--primary)!important;color:#FFF!important;border:none!important;border-radius:var(--radius)!important;padding:6px 16px!important;font-weight:600!important;box-shadow:none!important;}
.stButton>button:hover{ background:var(--primary-hover)!important; }
.element-container:has(.stDataFrame) .st-emotion-cache-1wmy9hl,
.element-container:has(.stDataEditor) .st-emotion-cache-1wmy9hl{
  background:#FFF;border:1px solid var(--border);border-radius:var(--radius);padding-top:6px;
}
button[role="tab"][aria-selected="true"]{ border-bottom:2px solid var(--primary)!important; color:var(--text)!important; }
section[data-testid="stSidebar"] .block-container{ background:var(--bg-main); border-right:1px solid var(--border); }
.pill{display:inline-block; padding:2px 8px; border-radius:999px; font-size:.8rem; border:1px solid #DDD; background:#F8FAFC}
.pill-pendente{ background:#FFF7ED; border-color:#FDBA74;}
.pill-nc{ background:#F3F4F6; border-color:#D1D5DB;}
.pill-enviado{ background:#EEF2FF; border-color:#C7D2FE;}
.pill-digitacao{ background:#ECFEFF; border-color:#BAE6FD;}
.pill-ok{ background:#ECFDF5; border-color:#A7F3D0;}
</style>
""", unsafe_allow_html=True)

def pill(situacao: str) -> str:
    s = (situacao or "").strip()
    cls = "pill"
    if s == "Pendente": cls += " pill-pendente"
    elif s == "Não Cobrar": cls += " pill-nc"
    elif s == "Enviado para pagamento": cls += " pill-enviado"
    elif s == "Aguardando Digitação - AMHP": cls += " pill-digitacao"
    elif s == "Finalizado": cls += " pill-ok"
    return f"<span class='{cls}'>{s or '-'}</span>"

def kpi_row(items, extra_class: str = ""):
    st.markdown(f"<div class='kpi-wrap {extra_class}'>", unsafe_allow_html=True)
    for it in items:
        st.markdown(
            f"""
<div class='kpi big'>
  <div class='label'>{it.get('label','')}</div>
  <div class='value'>{it.get('value','')}</div>
  { '<div class="hint">'+it.get('hint','')+'</div>' if it.get('hint') else '' }
</div>
""",
            unsafe_allow_html=True
        )
    st.markdown("</div>", unsafe_allow_html=True)

def app_header(title: str, subtitle: str = ""):
    st.markdown(
        f"""
<div class="app-header">
  <div class="title">🏥 {title}</div>
  <div class="sub">{subtitle}</div>
</div>
""",
        unsafe_allow_html=True
    )

def tab_header_with_home(title: str, home_label: str = "🏠 Início", btn_key_suffix: str = ""):
    col_t1, col_t2 = st.columns([8, 2])
    with col_t1:
        st.subheader(title)
    with col_t2:
        if st.button(home_label, key=f"btn_go_home_{btn_key_suffix}", use_container_width=True):
            st.session_state["goto_tab_label"] = "🏠 Início"
            st.session_state["__goto_nonce"] = st.session_state.get("__goto_nonce", 0) + 1
            st.rerun()

# ============================================================
# UTIL (datas, moeda)
# ============================================================
def _pt_date_to_dt(s):
    try:
        return datetime.strptime(str(s), "%d/%m/%Y").date()
    except Exception:
        try:
            return datetime.strptime(str(s), "%Y-%m-%d").date()
        except Exception:
            return None

def _to_ddmmyyyy(value):
    if value is None or value == "": return ""
    if isinstance(value, pd.Timestamp): return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime): return value.strftime("%d/%m/%Y")
    if isinstance(value, date): return value.strftime("%d/%m/%Y")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value), fmt).strftime("%d/%m/%Y")
        except Exception:
            pass
    return str(value)

def _to_float_or_none(v):
    if v is None or v == "": return None
    if isinstance(v, (int,float)): return float(v)
    s = re.sub(r"[^\d,\.\-]", "", str(v))
    if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
    elif "," in s:            s = s.replace(",", ".")
    try: return float(s)
    except: return None

def _format_currency_br(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)): return "R$ 0,00"
    try:
        v = float(v); s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {s}"
    except Exception:
        return f"R$ {v}"

# ============================================================
# UTIL — atendimento (normalização)
# ============================================================
def _att_norm(v) -> str:
    """
    Normaliza atendimento para comparação/armazenamento:
    - mantém apenas dígitos
    - remove zeros à esquerda
    - retorna '0' se ficar vazio
    """
    s = re.sub(r"\D", "", str(v or ""))
    s = s.lstrip("0")
    return s if s else "0"

def _att_to_number(v):
    """
    Converte atendimento para número (compatível com numero_internacao).
    Retorna None se não houver dígitos.
    """
    s = re.sub(r"\D", "", str(v or ""))
    if not s:
        return None
    try:
        return float(s)  # compatibilidade com schema atual (float)
    except Exception:
        return None

# ============================================================
# Helper de merge tolerante (evita KeyError com DF/coluna vazios)
# ============================================================
def _fmt_id_str(x):
    """
    Formata códigos numéricos (ex.: aviso, número de guia) como string sem '.0'.
    - Aceita None, str, int, float.
    - Remove espaços.
    - Converte floats inteiros (ex.: 6400413.0 -> '6400413').
    - Converte '385022.0' ou '3.85022e+05' em '385022'.
    - Mantém strings não-numéricas como vieram.
    """
    if x is None:
        return ""
    s = str(x).strip()
    if s == "":
        return ""
    try:
        f = float(s)
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
        return ("{0}".format(f)).replace(",", ".")
    except Exception:
        return s

def safe_merge(
    left: pd.DataFrame,
    right: pd.DataFrame,
    left_on: str,
    right_on: str,
    how: str = "left",
    suffixes=("", "_right"),
) -> pd.DataFrame:
    if not isinstance(left, pd.DataFrame) or left.empty:
        return left if isinstance(left, pd.DataFrame) else pd.DataFrame()
    if not isinstance(right, pd.DataFrame) or right.empty or (right_on not in right.columns):
        right = pd.DataFrame(columns=[right_on])
    if left_on not in left.columns:
        return left
    try:
        return left.merge(right, left_on=left_on, right_on=right_on, how=how, suffixes=suffixes)
    except KeyError:
        return left

# ============================
# BACKUP / RESTORE — Helpers
# ============================
SERVICE_KEY = st.secrets.get("SUPABASE_SERVICE_KEY", KEY)  # fallback no anon key
BUCKET = st.secrets.get("STORAGE_BACKUP_BUCKET", "backups")
admin_client: Client = create_client(URL, SERVICE_KEY)

def _fetch_all_rows(table: str, cols: str = "*", page_size: int = 1000, filters: Dict[str, Any] = None) -> List[Dict[str, Any]]:
    rows = []
    start = 0
    while True:
        q = supabase.table(table).select(cols).range(start, start + page_size - 1)
        if filters:
            for k, v in filters.items():
                q = q.eq(k, v)
        res = q.execute()
        chunk = res.data or []
        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        start += page_size
    return rows

def _to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")

def _now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def export_tables_to_zip(tables: List[str]) -> bytes:
    mem = _io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        meta = {
            "generated_at": datetime.now().isoformat(),
            "tables": tables,
            "app": "internacoes_supabase",
            "version": "v1"
        }
        zf.writestr("meta.json", json.dumps(meta, ensure_ascii=False, indent=2))
        for t in tables:
            data = _fetch_all_rows(t, "*")
            df = pd.DataFrame(data)
            zf.writestr(f"{t}.json", json.dumps(data, ensure_ascii=False, indent=2))
            zf.writestr(f"{t}.csv", _to_csv_bytes(df) if not df.empty else b"")
    return mem.getvalue()

def upload_zip_to_storage(zip_bytes: bytes, filename: str) -> bool:
    try:
        path = f"{filename}"
        admin_client.storage.from_(BUCKET).upload(path, zip_bytes, {"content-type": "application/zip", "upsert": True})
        return True
    except Exception as e:
        st.error(f"Falha ao enviar ao Storage: {e}")
        return False

def list_backups_from_storage(prefix: str = "", limit: int = 1000, offset: int = 0) -> list[dict]:
    try:
        options = {
            "limit": limit,
            "offset": offset,
            "sortBy": {"column": "updated_at", "order": "desc"}
        }
        res = admin_client.storage.from_(BUCKET).list(path=prefix or "", options=options)
        files = [
            f for f in res
            if isinstance(f, dict) and f.get("name", "").lower().endswith(".zip")
        ]
        def _get_when(x: dict):
            return x.get("updated_at") or x.get("last_modified") or x.get("created_at") or ""
        files.sort(key=_get_when, reverse=True)
        return files
    except Exception as e:
        st.error(f"Falha ao listar backups no Storage: {e}")
        return []

def download_backup_from_storage(name: str) -> bytes:
    try:
        return admin_client.storage.from_(BUCKET).download(name)
    except Exception as e:
        st.error(f"Falha no download do Storage: {e}")
        return b""

def _json_from_zip(zf: zipfile.ZipFile, name: str):
    try:
        with zf.open(name) as f:
            return json.loads(f.read().decode("utf-8"))
    except KeyError:
        return None

def restore_from_zip(zip_bytes: bytes, mode: str = "upsert") -> Dict[str, Any]:
    report = {"status": "ok", "details": []}
    try:
        with zipfile.ZipFile(_io.BytesIO(zip_bytes), mode="r") as zf:
            meta = _json_from_zip(zf, "meta.json") or {}
            tables = meta.get("tables") or ["hospitals", "internacoes", "procedimentos"]
            data_map = {}
            for t in tables:
                arr = _json_from_zip(zf, f"{t}.json")
                if arr is None:
                    try:
                        with zf.open(f"{t}.csv") as f:
                            df = pd.read_csv(f, dtype=str)
                            arr = json.loads(df.to_json(orient="records", force_ascii=False))
                    except Exception:
                        arr = []
                data_map[t] = arr or []

            ordered = ["hospitals", "internacoes", "procedimentos"]
            ordered = [t for t in ordered if t in data_map]

            if mode == "replace":
                for t in reversed(ordered):
                    try:
                        supabase.table(t).delete().neq("id", None).execute()
                        report["details"].append(f"{t}: apagado")
                    except APIError as e:
                        report["status"] = "error"
                        report["details"].append(f"{t}: falha ao apagar - {getattr(e,'message',e)}")
                        return report

            def _chunked_upsert(table: str, rows: List[Dict[str, Any]], chunk: int = 500):
                if not rows:
                    return 0
                total = 0
                for i in range(0, len(rows), chunk):
                    batch = rows[i:i+chunk]
                    try:
                        supabase.table(table).upsert(batch, on_conflict="id").execute()
                        total += len(batch)
                    except APIError as e:
                        try:
                            supabase.table(table).insert(batch).execute()
                            total += len(batch)
                        except APIError as e2:
                            report["status"] = "error"
                            report["details"].append(f"{table}: falha ao inserir/upsert - {getattr(e2,'message',e2)}")
                            break
                return total

            if "hospitals" in ordered:
                count = _chunked_upsert("hospitals", data_map["hospitals"])
                report["details"].append(f"hospitals: {count} registro(s) restaurado(s).")

            if "internacoes" in ordered:
                rows = data_map["internacoes"]
                for r in rows:
                    if "data_internacao" in r:
                        r["data_internacao"] = _to_ddmmyyyy(r["data_internacao"])
                    if "atendimento" in r:
                        r["atendimento"] = _att_norm(r["atendimento"])
                    if "numero_internacao" in r:
                        r["numero_internacao"] = _att_to_number(r["numero_internacao"])
                count = _chunked_upsert("internacoes", rows)
                report["details"].append(f"internacoes: {count} registro(s) restaurado(s).")

            if "procedimentos" in ordered:
                rows = data_map["procedimentos"]
                for r in rows:
                    if "data_procedimento" in r:
                        r["data_procedimento"] = _to_ddmmyyyy(r["data_procedimento"])
                    r["procedimento"] = r.get("procedimento") or "Cirurgia / Procedimento"
                    r["situacao"] = r.get("situacao") or "Pendente"
                    if "is_manual" in r:
                        r["is_manual"] = int(r["is_manual"] or 0)
                count = _chunked_upsert("procedimentos", rows)
                report["details"].append(f"procedimentos: {count} registro(s) restaurado(s).")

            invalidate_caches()
            return report
    except zipfile.BadZipFile:
        return {"status": "error", "details": ["Arquivo ZIP inválido."]}
    except Exception as e:
        return {"status": "error", "details": [f"Exceção: {e}"]}

# ============================================================
# CRUD — Supabase (tabelas minúsculas) + cache-aware
# ============================================================
@st.cache_data(ttl=TTL_LONG, show_spinner=False)
def get_hospitais(include_inactive: bool = False) -> list:
    try:
        query = supabase.table("hospitals").select("name, active")
        if not include_inactive:
            query = query.eq("active", 1)
        res = query.order("name").execute()
        return [r["name"] for r in (res.data or [])]
    except APIError as e:
        _sb_debug_error(e, "Falha ao buscar hospitais.")
        return []

def get_internacao_by_atendimento(att):
    """Busca por atendimento normalizado e, em fallback, por numero_internacao. NÃO cachear."""
    try:
        att_norm = _att_norm(att)
        res = supabase.table("internacoes").select("*").eq("atendimento", att_norm).execute()
        df = pd.DataFrame(res.data or [])
        if not df.empty:
            return df
        num = _att_to_number(att)
        if num is not None:
            res2 = supabase.table("internacoes").select("*").eq("numero_internacao", num).execute()
            return pd.DataFrame(res2.data or [])
        return pd.DataFrame()
    except APIError as e:
        _sb_debug_error(e, "Falha ao consultar internação.")
        return pd.DataFrame()

def criar_internacao(hospital, atendimento, paciente, data, convenio):
    att_norm = _att_norm(atendimento)
    num = _att_to_number(atendimento)
    payload = {
        "hospital": hospital,
        "atendimento": att_norm,
        "paciente": paciente,
        "data_internacao": _to_ddmmyyyy(data),
        "convenio": convenio,
        "numero_internacao": num
    }
    try:
        res = supabase.table("internacoes").insert(payload).execute()
        row = (res.data or [{}])[0]
        invalidate_caches()
        return int(row.get("id"))
    except APIError as e:
        _sb_debug_error(e, "Falha ao criar internação.")
        return None

def atualizar_internacao(internacao_id, **kwargs):
    update_data = {k: v for k, v in kwargs.items() if v is not None}
    if "data_internacao" in update_data:
        update_data["data_internacao"] = _to_ddmmyyyy(update_data["data_internacao"])
    try:
        supabase.table("internacoes").update(update_data).eq("id", int(internacao_id)).execute()
        invalidate_caches()
    except APIError as e:
        _sb_debug_error(e, "Falha ao atualizar internação.")

def deletar_internacao(internacao_id: int) -> bool:
    try:
        iid = int(internacao_id)
        pre_int = (
            supabase.table("internacoes")
            .select("id")
            .eq("id", iid)
            .limit(1)
            .execute()
        )
        if not (pre_int.data or []):
            st.info("A internação já não existe (nada a excluir).")
            return True

        pre_procs = (
            supabase.table("procedimentos")
            .select("id")
            .eq("internacao_id", iid)
            .execute()
        )
        qtd_procs = len(pre_procs.data or [])
        if qtd_procs > 0:
            supabase.table("procedimentos").delete().eq("internacao_id", iid).execute()

        chk_procs = (
            supabase.table("procedimentos")
            .select("id")
            .eq("internacao_id", iid)
            .limit(1)
            .execute()
        )
        if chk_procs.data:
            st.error("❌ Não foi possível excluir todos os procedimentos vinculados. Verifique RLS/Policies ou FKs.")
            return False

        supabase.table("internacoes").delete().eq("id", iid).execute()

        pos_int = (
            supabase.table("internacoes")
            .select("id")
            .eq("id", iid)
            .limit(1)
            .execute()
        )
        ok = len(pos_int.data or []) == 0
        if ok:
            invalidate_caches()
            return True
        else:
            st.error("❌ Não foi possível excluir a internação. Verifique RLS/Policies ou vínculos (FK).")
            return False
    except APIError as e:
        _sb_debug_error(e, "Falha ao deletar internação.")
        return False

def criar_procedimento(internacao_id, data_proc, profissional, procedimento,
                       situacao="Pendente", observacao=None, is_manual=0,
                       aviso=None, grau_participacao=None):
    payload = {
        "internacao_id": int(internacao_id),
        "data_procedimento": _to_ddmmyyyy(data_proc),
        "profissional": profissional,
        "procedimento": procedimento,
        "situacao": situacao or "Pendente",
        "observacao": observacao,
        "is_manual": int(is_manual or 0),
        "aviso": aviso,
        "grau_participacao": grau_participacao,
    }
    try:
        res = supabase.table("procedimentos").insert(payload).execute()
        data = res.data or []
        if not data:
            st.error("❌ O banco não confirmou a inclusão do procedimento (resposta vazia).")
            return None
        invalidate_caches()
        return int(data[0].get("id")) if data[0].get("id") is not None else True
    except APIError as e:
        _sb_debug_error(e, "Falha ao criar procedimento.")
        return None

def existe_procedimento_no_dia(internacao_id, data_proc):
    try:
        res = (
            supabase.table("procedimentos")
            .select("id")
            .eq("internacao_id", int(internacao_id))
            .eq("data_procedimento", _to_ddmmyyyy(data_proc))
            .eq("is_manual", 0)
            .limit(1)
            .execute()
        )
        return len(res.data or []) > 0
    except APIError as e:
        _sb_debug_error(e, "Falha ao verificar existência de procedimento no dia.")
        return False

def atualizar_procedimento(proc_id, procedimento=None, situacao=None,
                           observacao=None, grau_participacao=None, aviso=None):
    update_data = {}
    if procedimento is not None: update_data["procedimento"] = procedimento
    if situacao is not None: update_data["situacao"] = situacao
    if observacao is not None: update_data["observacao"] = observacao
    if grau_participacao is not None: update_data["grau_participacao"] = grau_participacao
    if aviso is not None: update_data["aviso"] = aviso
    if not update_data: return
    try:
        supabase.table("procedimentos").update(update_data).eq("id", int(proc_id)).execute()
        invalidate_caches()
    except APIError as e:
        _sb_debug_error(e, "Falha ao atualizar procedimento.")

def deletar_procedimento(proc_id: int) -> bool:
    try:
        pre = (
            supabase.table("procedimentos")
            .select("id")
            .eq("id", int(proc_id))
            .limit(1)
            .execute()
        )
        if not (pre.data or []):
            st.info("Registro já não existe (nada a excluir).")
            return True

        supabase.table("procedimentos").delete().eq("id", int(proc_id)).execute()

        pos = (
            supabase.table("procedimentos")
            .select("id")
            .eq("id", int(proc_id))
            .limit(1)
            .execute()
        )
        ok = len(pos.data or []) == 0
        if ok:
            invalidate_caches()
            return True
        else:
            st.error("❌ Não foi possível excluir. Verifique RLS/Policies ou vínculos (FK).")
            return False
    except APIError as e:
        _sb_debug_error(e, "Falha ao deletar procedimento.")
        return False

def quitar_procedimento(proc_id, data_quitacao=None, guia_amhptiss=None, valor_amhptiss=None,
                        guia_complemento=None, valor_complemento=None, quitacao_observacao=None):
    update_data = {
        "quitacao_data": _to_ddmmyyyy(data_quitacao) if data_quitacao else None,
        "quitacao_guia_amhptiss": (_fmt_id_str(guia_amhptiss) or None),
        "quitacao_valor_amhptiss": valor_amhptiss,
        "quitacao_guia_complemento": (_fmt_id_str(guia_complemento) or None),
        "quitacao_valor_complemento": valor_complemento,
        "quitacao_observacao": quitacao_observacao,
        "situacao": "Finalizado",
    }
    update_data = {k:v for k,v in update_data.items() if v is not None or k=="situacao"}
    try:
        supabase.table("procedimentos").update(update_data).eq("id", int(proc_id)).execute()
        invalidate_caches()
    except APIError as e:
        _sb_debug_error(e, "Falha ao quitar procedimento.")

def _excel_quitacoes_colunas_fixas(df: pd.DataFrame) -> bytes:
    if df is None or df.empty:
        return b""
    cols_pdf = [
        "quitacao_data","hospital","atendimento","paciente","convenio",
        "profissional","grau_participacao","data_procedimento",
        "quitacao_guia_amhptiss","quitacao_valor_amhptiss",
        "quitacao_guia_complemento","quitacao_valor_complemento",
    ]
    base = df.copy()
    for c in cols_pdf:
        if c not in base.columns:
            base[c] = ""
    for col in ["quitacao_guia_amhptiss","quitacao_guia_complemento"]:
        if col in base.columns:
            base[col] = base[col].apply(_fmt_id_str)

    def _to_date_or_none(s):
        d = _pt_date_to_dt(s)
        return pd.to_datetime(d) if d else pd.NaT
    base["quitacao_data_x"] = base["quitacao_data"].apply(_to_date_or_none)
    base["data_procedimento_x"] = base["data_procedimento"].apply(_to_date_or_none)

    base["quitacao_valor_amhptiss_x"] = pd.to_numeric(base["quitacao_valor_amhptiss"], errors="coerce")
    base["quitacao_valor_complemento_x"] = pd.to_numeric(base["quitacao_valor_complemento"], errors="coerce")

    out = pd.DataFrame({
        "Quitação": base["quitacao_data_x"],
        "Hospital": base["hospital"],
        "Atendimento": base["atendimento"],
        "Paciente": base["paciente"],
        "Convênio": base["convenio"],
        "Profissional": base["profissional"],
        "Grau": base["grau_participacao"],
        "Data Proc.": base["data_procedimento_x"],
        "Guia AMHPTISS": base["quitacao_guia_amhptiss"],
        "R$ AMHPTISS": base["quitacao_valor_amhptiss_x"],
        "Guia Compl.": base["quitacao_guia_complemento"],
        "R$ Compl.": base["quitacao_valor_complemento_x"],
    })

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = "Quitações"
        out.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import numbers, Alignment, Font
        date_fmt = "dd/mm/yyyy"
        money_fmt = u'[$R$-pt_BR] #,##0.00'

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_quit = headers.index("Quitação") + 1
        idx_proc = headers.index("Data Proc.") + 1
        idx_v1 = headers.index("R$ AMHPTISS") + 1
        idx_v2 = headers.index("R$ Compl.") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_quit, max_col=idx_quit):
            for cell in row:
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_proc, max_col=idx_proc):
            for cell in row:
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_v1, max_col=idx_v1):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx_v2, max_col=idx_v2):
            for cell in row:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

        for col_title in ["Atendimento","Guia AMHPTISS","Guia Compl."]:
            cidx = headers.index(col_title) + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=cidx, max_col=cidx):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        col_width_map = {
            "Quitação": 11,
            "Hospital": 16,
            "Atendimento": 12,
            "Paciente": 32,
            "Convênio": 18,
            "Profissional": 22,
            "Grau": 12,
            "Data Proc.": 11,
            "Guia AMHPTISS": 16,
            "R$ AMHPTISS": 14,
            "Guia Compl.": 16,
            "R$ Compl.": 14,
        }
        for col_cells in ws.iter_cols(min_row=1, max_row=1):
            title = col_cells[0].value
            if title in col_width_map:
                ws.column_dimensions[col_cells[0].column_letter].width = col_width_map[title]

        ws.freeze_panes = "A2"

    return buf.getvalue()

def reverter_quitacao(proc_id: int):
    update_data = {
        "quitacao_data": None,
        "quitacao_guia_amhptiss": None,
        "quitacao_valor_amhptiss": None,
        "quitacao_guia_complemento": None,
        "quitacao_valor_complemento": None,
        "quitacao_observacao": None,
        "situacao": "Enviado para pagamento",
    }
    try:
        supabase.table("procedimentos").update(update_data).eq("id", int(proc_id)).execute()
        invalidate_caches()
    except APIError as e:
        _sb_debug_error(e, "Falha ao reverter quitação.")

@st.cache_data(ttl=TTL_SHORT, show_spinner=False)
def get_procedimentos(internacao_id):
    try:
        res = supabase.table("procedimentos").select("*").eq("internacao_id", int(internacao_id)).execute()
        return pd.DataFrame(res.data or [])
    except APIError as e:
        _sb_debug_error(e, "Falha ao listar procedimentos.")
        return pd.DataFrame()

@st.cache_data(ttl=TTL_SHORT, show_spinner=False)
def get_quitacao_by_proc_id(proc_id: int):
    try:
        r1 = supabase.table("procedimentos").select("*").eq("id", int(proc_id)).limit(1).execute()
        dfp = pd.DataFrame(r1.data or [])
        if dfp.empty: return dfp
        iid = int(dfp["internacao_id"].iloc[0])
        r2 = supabase.table("internacoes").select("*").eq("id", iid).limit(1).execute()
        dfi = pd.DataFrame(r2.data or [])
        if dfi.empty: return dfp
        df = safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left", suffixes=("", "_int"))
        return df
    except APIError as e:
        _sb_debug_error(e, "Falha ao consultar quitação.")
        return pd.DataFrame()

# ============================================================
# Consultas cacheadas (bases usadas em telas pesadas)
# Agora com opção de usar VIEW (USE_DB_VIEW) e fallback para merge local
# ============================================================
@st.cache_data(ttl=TTL_MED, show_spinner=False)
def _home_fetch_base_df() -> pd.DataFrame:
    if USE_DB_VIEW:
        try:
            res = supabase.table("vw_procedimentos_internacoes").select(
                "procedimento_id, internacao_id, data_procedimento, procedimento, profissional, situacao, aviso, grau_participacao, "
                "atendimento, paciente, hospital, convenio, data_internacao"
            ).execute()
            df = pd.DataFrame(res.data or [])
            if "procedimento_id" in df.columns and "id" not in df.columns:
                df = df.rename(columns={"procedimento_id": "id"})
            return df
        except APIError as e:
            _sb_debug_error(e, "Falha na view vw_procedimentos_internacoes. Usando fallback local.")
    try:
        res_p = supabase.table("procedimentos").select(
            "id, internacao_id, data_procedimento, procedimento, profissional, situacao, aviso, grau_participacao"
        ).execute()
        df_p = pd.DataFrame(res_p.data or [])
        if df_p.empty:
            return pd.DataFrame(columns=[
                "internacao_id","atendimento","paciente","hospital","convenio","data_internacao",
                "id","data_procedimento","procedimento","profissional","situacao","aviso","grau_participacao"
            ])
        ids = sorted(set(int(x) for x in df_p["internacao_id"].dropna().tolist()))
        res_i = supabase.table("internacoes").select(
            "id, atendimento, paciente, hospital, convenio, data_internacao"
        ).in_("id", ids).execute() if ids else None
        df_i = pd.DataFrame(res_i.data or []) if res_i else pd.DataFrame()
        return safe_merge(
            df_p,
            df_i[["id", "atendimento", "paciente", "hospital", "convenio", "data_internacao"]] if not df_i.empty else df_i,
            left_on="internacao_id",
            right_on="id",
            how="left",
            suffixes=("", "_int"),
        )
    except APIError as e:
        _sb_debug_error(e, "Falha ao carregar dados para a Home.")
        return pd.DataFrame()

@st.cache_data(ttl=TTL_MED, show_spinner=False)
def _listar_profissionais_cache() -> list:
    try:
        res_dist = supabase.table("procedimentos").select("profissional").execute()
        df_pros = pd.DataFrame(res_dist.data or [])
        if "profissional" in df_pros.columns:
            lista_profissionais = sorted({
                str(x).strip() for x in df_pros["profissional"].dropna()
                if str(x).strip()
            })
        else:
            lista_profissionais = []
        return lista_profissionais
    except APIError:
        return []

@st.cache_data(ttl=TTL_MED, show_spinner=False)
def _rel_cirurgias_base_df() -> pd.DataFrame:
    if USE_DB_VIEW:
        try:
            res = supabase.table("vw_procedimentos_internacoes").select(
                "procedimento_id, internacao_id, data_procedimento, aviso, profissional, procedimento, grau_participacao, situacao, "
                "hospital, atendimento, paciente, convenio"
            ).eq("procedimento", "Cirurgia / Procedimento").execute()
            df = pd.DataFrame(res.data or [])
            if "procedimento_id" in df.columns and "id" not in df.columns:
                df = df.rename(columns={"procedimento_id": "id"})
            return df
        except APIError as e:
            _sb_debug_error(e, "Falha na view (rel cirurgias). Usando fallback local.")
    try:
        resp = supabase.table("procedimentos").select(
            "internacao_id, data_procedimento, aviso, profissional, procedimento, grau_participacao, situacao"
        ).eq("procedimento", "Cirurgia / Procedimento").execute()
        dfp = pd.DataFrame(resp.data or [])
        if dfp.empty:
            return pd.DataFrame()
        ids = sorted(set(int(x) for x in dfp["internacao_id"].dropna().tolist()))
        if ids:
            resi = supabase.table("internacoes").select(
                "id, hospital, atendimento, paciente, convenio"
            ).in_("id", ids).execute()
            dfi = pd.DataFrame(resi.data or [])
        else:
            dfi = pd.DataFrame(columns=["id","hospital","atendimento","paciente","convenio"])
        return safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left")
    except APIError as e:
        _sb_debug_error(e, "Falha ao carregar dados para Relatório.")
        return pd.DataFrame()

@st.cache_data(ttl=TTL_MED, show_spinner=False)
def _rel_quitacoes_base_df() -> pd.DataFrame:
    if USE_DB_VIEW:
        try:
            res = supabase.table("vw_procedimentos_internacoes").select(
                "procedimento_id, internacao_id, data_procedimento, profissional, grau_participacao, situacao, "
                "quitacao_data, quitacao_guia_amhptiss, quitacao_guia_complemento, "
                "quitacao_valor_amhptiss, quitacao_valor_complemento, "
                "hospital, atendimento, paciente, convenio"
            ).not_.is_("quitacao_data", None).eq("procedimento", "Cirurgia / Procedimento").execute()
            df = pd.DataFrame(res.data or [])
            if "procedimento_id" in df.columns and "id" not in df.columns:
                df = df.rename(columns={"procedimento_id": "id"})
            return df
        except APIError as e:
            _sb_debug_error(e, "Falha na view (rel quitações). Usando fallback local.")
    try:
        resp = supabase.table("procedimentos").select(
            "internacao_id, data_procedimento, profissional, grau_participacao, situacao, "
            "quitacao_data, quitacao_guia_amhptiss, quitacao_guia_complemento, "
            "quitacao_valor_amhptiss, quitacao_valor_complemento"
        ).eq("procedimento", "Cirurgia / Procedimento").not_.is_("quitacao_data", None).execute()
        dfp = pd.DataFrame(resp.data or [])
        if dfp.empty:
            return pd.DataFrame()
        ids = sorted(set(int(x) for x in dfp["internacao_id"].dropna().tolist()))
        if ids:
            resi = supabase.table("internacoes").select(
                "id, hospital, atendimento, paciente, convenio"
            ).in_("id", ids).execute()
            dfi = pd.DataFrame(resi.data or [])
        else:
            dfi = pd.DataFrame()
        return safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left")
    except APIError as e:
        _sb_debug_error(e, "Falha ao carregar dados de quitações.")
        return pd.DataFrame()

@st.cache_data(ttl=TTL_MED, show_spinner=False)
def _quitacao_pendentes_base_df() -> pd.DataFrame:
    if USE_DB_VIEW:
        try:
            res = supabase.table("vw_procedimentos_internacoes").select(
                "procedimento_id, internacao_id, data_procedimento, profissional, aviso, situacao, "
                "quitacao_data, quitacao_guia_amhptiss, quitacao_valor_amhptiss, "
                "quitacao_guia_complemento, quitacao_valor_complemento, quitacao_observacao, "
                "hospital, atendimento, paciente, convenio"
            ).eq("procedimento", "Cirurgia / Procedimento").eq("situacao", "Enviado para pagamento").execute()
            df = pd.DataFrame(res.data or [])
            if "procedimento_id" in df.columns and "id" not in df.columns:
                df = df.rename(columns={"procedimento_id": "id"})
            return df
        except APIError as e:
            _sb_debug_error(e, "Falha na view (pendências quitação). Usando fallback local.")
    try:
        resp = supabase.table("procedimentos").select(
            "id, internacao_id, data_procedimento, profissional, aviso, situacao, "
            "quitacao_data, quitacao_guia_amhptiss, quitacao_valor_amhptiss, "
            "quitacao_guia_complemento, quitacao_valor_complemento, quitacao_observacao"
        ).eq("procedimento", "Cirurgia / Procedimento").eq("situacao", "Enviado para pagamento").execute()
        dfp = pd.DataFrame(resp.data or [])
        if dfp.empty:
            return pd.DataFrame()
        ids = sorted(set(int(x) for x in dfp["internacao_id"].dropna().tolist()))
        if ids:
            resi = supabase.table("internacoes").select("id, hospital, atendimento, paciente, convenio").in_("id", ids).execute()
            dfi = pd.DataFrame(resi.data or [])
        else:
            dfi = pd.DataFrame()
        return safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left", suffixes=("", "_int"))
    except APIError as e:
        _sb_debug_error(e, "Falha ao carregar pendências de quitação.")
        return pd.DataFrame()

# ============================================================
# INICIALIZAÇÃO UI
# ============================================================
st.set_page_config(page_title="Gestão de Internações", page_icon="🏥", layout="wide")
inject_css()
app_header("Sistema de Internações — Supabase",
           "Importação, edição, quitação e relatórios (banco em nuvem)")

def _switch_to_tab_by_label(tab_label: str):
    nonce = int(st.session_state.get("__goto_nonce", 0))
    js = """
<script>
// nonce: __NONCE__
(function(){
  const target = __TAB_LABEL__;
  const norm = (s)=> (s||"").replace(/\s+/g, " ").trim();
  let attempts = 0;
  const maxAttempts = 20;
  const timer = setInterval(()=>{
    attempts++;
    const tabs = window.parent.document.querySelectorAll('button[role="tab"]');
    for (const t of tabs) {
      const txt = norm(t.textContent || t.innerText);
      if (txt.includes(norm(target))) {
        t.click();
        clearInterval(timer);
        return;
      }
    }
    if (attempts >= maxAttempts) {
      clearInterval(timer);
      console.warn("Tab não encontrada para:", target);
    }
  }, 100);
})();
</script>
"""
    js = js.replace("__TAB_LABEL__", json.dumps(tab_label))
    js = js.replace("__NONCE__", str(nonce))
    components.html(js, height=0, width=0)

tabs = st.tabs([
    "🏠 Início",
    "📤 Importar Arquivo",
    "🔍 Consultar Internação",
    "📑 Relatórios",
    "💼 Quitação",
    "⚙️ Sistema",
])

# ============================================================
# 🏠 0) INÍCIO
# ============================================================
with tabs[0]:
    st.subheader("🏠 Tela Inicial")
    if "home_status" not in st.session_state:
        st.session_state["home_status"] = None
    hoje = date.today()
    ini_mes = hoje.replace(day=1)

    colf1, colf2 = st.columns([2,3])
    with colf1:
        filtro_hosp_home = st.selectbox("Hospital", ["Todos"] + get_hospitais(), index=0, key="home_f_hosp")
    with colf2:
        st.write(" ")
        st.caption("Períodos (opcionais)")
        cbox1, cbox2 = st.columns(2)
        with cbox1:
            use_int_range = st.checkbox("Filtrar por data da internação", key="home_use_int_range", value=False)
        with cbox2:
            use_proc_range = st.checkbox("Filtrar por data do procedimento", key="home_use_proc_range", value=False)
        if use_int_range or use_proc_range:
            cold1, cold2, cold3, cold4 = st.columns(4)
            with cold1:
                int_ini = st.date_input("Internação — início", value=st.session_state.get("home_f_int_ini", ini_mes), key="home_f_int_ini")
            with cold2:
                int_fim = st.date_input("Internação — fim", value=st.session_state.get("home_f_int_fim", hoje), key="home_f_int_fim")
            with cold3:
                proc_ini = st.date_input("Procedimento — início", value=st.session_state.get("home_f_proc_ini", ini_mes), key="home_f_proc_ini")
            with cold4:
                proc_fim = st.date_input("Procedimento — fim", value=st.session_state.get("home_f_proc_fim", hoje), key="home_f_proc_fim")

    df_all = _home_fetch_base_df()

    if df_all.empty:
        df_f = df_all.copy()
    else:
        def _safe_pt_date(s):
            try:
                return datetime.strptime(str(s).strip(), "%d/%m/%Y").date()
            except Exception:
                try:
                    return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
                except Exception:
                    return None
        df_all["_int_dt"] = df_all["data_internacao"].apply(_safe_pt_date)
        df_all["_proc_dt"] = df_all["data_procedimento"].apply(_safe_pt_date)
        mask = pd.Series([True]*len(df_all), index=df_all.index)
        if filtro_hosp_home != "Todos":
            mask &= (df_all["hospital"] == filtro_hosp_home)
        if use_int_range:
            mask &= df_all["_int_dt"].notna()
            mask &= (df_all["_int_dt"] >= st.session_state["home_f_int_ini"])
            mask &= (df_all["_int_dt"] <= st.session_state["home_f_int_fim"])
        if use_proc_range:
            mask &= df_all["_proc_dt"].notna()
            mask &= (df_all["_proc_dt"] >= st.session_state["home_f_proc_ini"])
            mask &= (df_all["_proc_dt"] <= st.session_state["home_f_proc_fim"])
        df_f = df_all[mask].copy()

    def _count_status(df: pd.DataFrame, status: str) -> int:
        if df is None or df.empty:
            return 0
        col = "situacao" if "situacao" in df.columns else None
        if col is None:
            return 0
        return int((df[col] == status).sum())

    tot_pendente = _count_status(df_f, "Pendente")
    tot_finalizado = _count_status(df_f, "Finalizado")
    tot_nao_cobrar = _count_status(df_f, "Não Cobrar")

    def _toggle_home_status(target: str):
        curr = st.session_state.get("home_status")
        st.session_state["home_status"] = None if curr == target else target
        st.rerun()

    active = st.session_state.get("home_status")
    c1, c2, c3 = st.columns(3)
    with c1:
        kpi_row([{"label":"Pendentes", "value": f"{tot_pendente}", "hint": "Todos os procedimentos"}], extra_class="center")
        lbl = "🔽 Esconder Pendentes" if active == "Pendente" else "👁️ Ver Pendentes"
        st.markdown("<div class='kpi-action'>", unsafe_allow_html=True)
        if st.button(lbl, key="kpi_btn_pend", use_container_width=True):
            _toggle_home_status("Pendente")
        st.markdown("</div>", unsafe_allow_html=True)
    with c2:
        kpi_row([{"label":"Finalizadas", "value": f"{tot_finalizado}", "hint": "Todos os procedimentos"}], extra_class="center")
        lbl = "🔽 Esconder Finalizadas" if active == "Finalizado" else "👁️ Ver Finalizadas"
        st.markdown("<div class='kpi-action'>", unsafe_allow_html=True)
        if st.button(lbl, key="kpi_btn_fin", use_container_width=True):
            _toggle_home_status("Finalizado")
        st.markdown("</div>", unsafe_allow_html=True)
    with c3:
        kpi_row([{"label":"Não Cobrar", "value": f"{tot_nao_cobrar}", "hint": "Todos os procedimentos"}], extra_class="center")
        lbl = "🔽 Esconder Não Cobrar" if active == "Não Cobrar" else "👁️ Ver Não Cobrar"
        st.markdown("<div class='kpi-action'>", unsafe_allow_html=True)
        if st.button(lbl, key="kpi_btn_nc", use_container_width=True):
            _toggle_home_status("Não Cobrar")
        st.markdown("</div>", unsafe_allow_html=True)

    status_sel_home = st.session_state.get("home_status")
    if status_sel_home:
        st.divider()
        st.subheader(f"📋 Internações com ao menos 1 procedimento em: **{status_sel_home}**")
        cc1, _ = st.columns([1, 6])
        with cc1:
            if st.button("Fechar lista", key="btn_close_list", type="secondary", use_container_width=True):
                st.session_state["home_status"] = None
                st.rerun()

        if df_f.empty:
            st.info("Nenhuma internação encontrada com os filtros aplicados.")
        else:
            df_status = df_f[df_f["situacao"] == status_sel_home].copy()
            if df_status.empty:
                st.info("Nenhuma internação encontrada para este status com os filtros atuais.")
            else:
                cols_show = ["internacao_id","atendimento","paciente","hospital","convenio","data_internacao"]
                df_ints = df_status[cols_show].drop_duplicates(subset=["internacao_id"]).copy()

                def _safe_pt_date_int(s):
                    try:
                        return datetime.strptime(str(s).strip(), "%d/%m/%Y").date()
                    except Exception:
                        try:
                            return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
                        except Exception:
                            return None
                df_ints["_int_dt"] = df_ints["data_internacao"].apply(_safe_pt_date_int)
                df_ints = (
                    df_ints.sort_values(by=["_int_dt","hospital","paciente"], ascending=[False,True,True])
                    .drop(columns=["_int_dt"])
                )
                for _, r in df_ints.iterrows():
                    i1, i2, i3, i4 = st.columns([3, 3, 3, 2])
                    with i1:
                        st.markdown(f"**Atendimento:** {r['atendimento']}  \n**Paciente:** {r.get('paciente') or '-'}")
                    with i2:
                        st.markdown(f"**Hospital:** {r.get('hospital') or '-'}  \n**Convênio:** {r.get('convenio') or '-'}")
                    with i3:
                        st.markdown(f"**Data internação:** {r.get('data_internacao') or '-'}")
                    with i4:
                        if st.button("🔎 Abrir na Consulta", key=f"open_cons_{int(r['internacao_id'])}", use_container_width=True):
                            st.session_state["consulta_codigo"] = str(r["atendimento"])
                            st.session_state["goto_tab_label"] = "🔍 Consultar Internação"
                if st.session_state.get("consulta_codigo"):
                    st.caption(f"🔎 Atendimento **{st.session_state['consulta_codigo']}** pronto para consulta na aba **'🔍 Consultar Internação'**.")

# ============================================================
# 📤 1) IMPORTAR (Importação primeiro, cadastro manual depois)
# ============================================================
with tabs[1]:
    tab_header_with_home("📤 Importar arquivo", btn_key_suffix="import")

    # --- Seção: Importação de CSV ---
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    hospitais = get_hospitais()
    hospital = st.selectbox("Hospital para esta importação:", hospitais, key="import_csv_hospital")
    arquivo = st.file_uploader("Selecione o arquivo CSV", key="import_csv_uploader")

    if parse_tiss_original is None:
        st.info("Adicione o arquivo parser.py com a função parse_tiss_original() para habilitar a importação.")
    elif arquivo:
        raw_bytes = arquivo.getvalue()
        try:
            csv_text = raw_bytes.decode("latin1")
        except UnicodeDecodeError:
            csv_text = raw_bytes.decode("utf-8-sig", errors="ignore")

        # --- Lê registros do parser ---
        registros = parse_tiss_original(csv_text)
              
        st.success(f"{len(registros)} registros interpretados!")

        # Profissionais distintos encontrados no arquivo
        pros = sorted({(r.get("profissional") or "").strip() for r in registros if r.get("profissional")})

        # --- KPI por (atendimento, aviso) ---
        pares_aviso = sorted({
            ((r.get("atendimento") or "").strip(), (_fmt_id_str(r.get("aviso")) or ""))
            for r in registros if r.get("atendimento")
        })
        kpi_row([
            {"label": "Registros no arquivo", "value": f"{len(registros):,}".replace(",", ".")},
            {"label": "Médicos distintos", "value": f"{len(pros):,}".replace(",", ".")},
            {"label": "Pares (atendimento, aviso)", "value": f"{len([p for p in pares_aviso if p[0] and p[1]]):,}".replace(",", ".")},
        ])

        # --- Seleção de médicos ---
        st.subheader("👨‍⚕️ Seleção de médicos")
        if "import_all_docs" not in st.session_state: st.session_state["import_all_docs"] = True
        if "import_selected_docs" not in st.session_state: st.session_state["import_selected_docs"] = []
        colsel1, colsel2 = st.columns([1, 3])
        with colsel1:
            import_all = st.checkbox("Importar todos os médicos", value=st.session_state["import_all_docs"], key="import_all_docs_chk")
        with colsel2:
            if import_all:
                st.info("Todos os médicos do arquivo serão importados.")
                selected_pros = pros[:]
            else:
                default_pre = sorted([p for p in pros if p in ALWAYS_SELECTED_PROS])
                selected_pros = st.multiselect(
                    "Médicos a importar (os da lista fixa sempre serão incluídos na gravação):",
                    options=pros,
                    default=st.session_state["import_selected_docs"] or default_pre,
                    key="import_selected_docs_ms"
                )
        st.session_state["import_all_docs"] = import_all
        st.session_state["import_selected_docs"] = selected_pros
        always_in_file = [p for p in pros if p in ALWAYS_SELECTED_PROS]
        final_pros = sorted(set(selected_pros if not import_all else pros).union(ALWAYS_SELECTED_PROS))
        st.caption(f"Médicos fixos (sempre incluídos, quando presentes): {', '.join(sorted(ALWAYS_SELECTED_PROS))}")
        st.info(f"Médicos considerados: {', '.join(final_pros) if final_pros else '(nenhum)'}")

        
        # --- (1) Agrupa por (atendimento, aviso) preservando a ordem do CSV
        grupos = OrderedDict()
        for r in registros:
            att = (r.get("atendimento") or "").strip()
            aviso_fmt = _fmt_id_str(r.get("aviso"))
            if not att or not aviso_fmt:
                continue
            key = (att, aviso_fmt)
            grupos.setdefault(key, []).append(r)

        def _escolher_profissional(rows: list[dict]) -> tuple[str, str]:
            """
            Decide o 'prof_escolhido' para um grupo (atendimento, aviso).
        
            Regra A (prioridade 1):
              - Se a linha MESTRE tem 'profissional' (campo parseado) NÃO vazio -> usa este (regra = "A").
        
            Regra B (prioridade 2) — NOVA:
              - Aplica somente se a MESTRE NÃO tem profissional (campo vazio).
              - Procura o PRIMEIRO 'profissional' NÃO vazio nas LINHAS-FILHAS (ordem do CSV) -> usa este (regra = "B").
        
            Fallback (prioridade 3):
              - Se nada acima, tenta varrer as células (__cells__) da MESTRE e depois das FILHAS
                por algum nome de 'ALWAYS_SELECTED_PROS' -> se achar, retorna ("A" p/ mestre, "B" p/ filha).
              - Se ainda assim não achar, retorna ("", "SKIP").
            """
            def _cells_of(r: dict) -> list[str]:
                if "__cells__" in r and isinstance(r["__cells__"], list):
                    return r["__cells__"]
                # fallback para parser sem __cells__
                return [
                    r.get("procedimento", ""),
                    r.get("convenio", ""),
                    r.get("profissional", ""),
                    r.get("anestesista", ""),
                    r.get("tipo", ""),
                    r.get("quarto", ""),
                ]
        
            # ---------- Regra A: mestre com profissional ----------
            prof_mestre = (rows[0].get("profissional") or "").strip()
            if prof_mestre:
                return prof_mestre, "A"
        
            # ---------- Regra B: mestre sem profissional -> 1º das FILHAS ----------
            for rr in rows[1:]:
                prof = (rr.get("profissional") or "").strip()
                if prof:
                    return prof, "B"
        
            # ---------- Fallback: varredura por nomes fixos nas células ----------
            # (a) mestre por células
            name = find_allowed_in_row(_cells_of(rows[0]))
            if name:
                return name, "A"
        
            # (b) filhas por células (ordem CSV)
            for rr in rows[1:]:
                name = find_allowed_in_row(_cells_of(rr))
                if name:
                    return name, "B"
        
            # Nada encontrado
            return "", "SKIP"

        # --- Debug 1: Contagens e SKIPs
        total_grupos = len(grupos)
        total_AB = sum(1 for g in grupos_info if g["regra"] in ("A","B"))
        total_SKIP = sum(1 for g in grupos_info if g["regra"] == "SKIP")
        
        with st.expander("🔎 A/B vs SKIP (debug temporário)"):
            st.write("Grupos totais:", total_grupos)
            st.write("Grupos com Regra A/B:", total_AB)
            st.write("Grupos SKIP:", total_SKIP)
            if total_SKIP:
                st.table(pd.DataFrame([
                    {
                        "atendimento": g["atendimento"],
                        "aviso": g["aviso"],
                        "profissionais_parser": [ (r.get("profissional") or "") for r in g["rows"] ][:3],  # amostra
                    }
                    for g in grupos_info if g["regra"] == "SKIP"
                ]))        

        
        # --- Debug 2: Diferença entre pares vindos do arquivo e os consolidados (A/B)
        pairs_from_file = {
            ((r.get("atendimento") or "").strip(), (_fmt_id_str(r.get("aviso")) or ""))
            for r in registros if r.get("atendimento")
        }
        pairs_AB = { (g["atendimento"], g["aviso"]) for g in grupos_info if g["regra"] in ("A","B") }
        
        faltando_em_AB = sorted(pairs_from_file - pairs_AB)
        a_mais_em_AB = sorted(pairs_AB - pairs_from_file)
        
        with st.expander("🧮 Diff de pares (arquivo × A/B)"):
            st.write(f"Pares no arquivo: {len(pairs_from_file)} | Pares A/B: {len(pairs_AB)}")
            st.write("Faltando em A/B:", faltando_em_AB)
            st.write("A mais em A/B:", a_mais_em_AB)
            # Se quiser detalhar o grupo faltante:
            for att, av in faltando_em_AB[:3]:
                rows = grupos.get((att, av), [])
                st.write(f"Grupo faltante -> atendimento={att}, aviso={av}")
                if rows:
                    st.write("Profissionais (parser):", [r.get("profissional") for r in rows])
                    st.write("Células (mestre):", rows[0].get("__cells__"))

        
        # --- debug 3 Auto-detector do 24º par perdido 
        pairs_from_file = {
            ((r.get("atendimento") or "").strip(), (_fmt_id_str(r.get("aviso")) or ""))
            for r in registros if r.get("atendimento")
        }
        pairs_AB = { (g["atendimento"], g["aviso"]) for g in grupos_info if g["regra"] in ("A","B") }
        
        faltando_em_AB = sorted(pairs_from_file - pairs_AB)
        pairs_considerados = set()  # será preenchido abaixo, depois que você montar grupos_considerados


        # --- Debug 3: Pares considerados para gravação após filtro de médicos
        import_all = st.session_state.get("import_all_docs", True)
        final_pros_set = set(final_pros)  # já definido acima
        
        grupos_considerados = [
            g for g in grupos_info
            if g["regra"] in ("A","B")
            and g["prof_escolhido"]
            and (import_all or g["prof_escolhido"] in final_pros_set)
        ]
        
        with st.expander("✅ Pares considerados (após filtro de médicos)"):
            st.write("import_all:", import_all)
            st.write("Médicos considerados:", sorted(final_pros_set))
            st.write("Pares (atendimento,aviso) considerados:", len(grupos_considerados))
            st.table(pd.DataFrame([
                {"atendimento": g["atendimento"], "aviso": g["aviso"], "prof": g["prof_escolhido"], "regra": g["regra"]}
                for g in grupos_considerados
            ]))

        
             
        # ⚠️ calcular pares_considerados SOMENTE DEPOIS de tudo pronto
        pairs_considerados = set()
        for g in grupos_considerados:
            att = g["atendimento"]
            av  = g["aviso"]
            if att and av:
                pairs_considerados.add((att, av))

        
        # 1) Se faltou antes (na Regra A/B)
        if faltando_em_AB:
            att, av = faltando_em_AB[0]
            st.warning(f"⚠️ Par FALTANDO na Regra A/B: atendimento={att}, aviso={av}")
            rows = grupos.get((att, av), [])
            if not rows:
                st.write("Grupo não encontrado no dicionário 'grupos' (verifique chaves).")
            else:
                st.write("• Profissionais parseados por linha do grupo:", [ (r.get('profissional') or "") for r in rows ])
                st.write("• CÉLULAS da linha-mestre (para Regra A):", rows[0].get("__cells__"))
                # Se quiser ver todas as linhas do grupo (mestre + filhas):
                for i, rr in enumerate(rows):
                    st.caption(f"— Linha {i} (mestre=0):")
                    st.write(rr.get("__cells__"))
        else:
            # 2) Se A/B tem 24, veja se o filtro de médicos derrubou algum
            faltando_no_filtro = sorted(pairs_AB - pairs_considerados)
            if faltando_no_filtro:
                att, av = faltando_no_filtro[0]
                st.warning(f"⚠️ Par FALTANDO APÓS FILTRO DE MÉDICOS: atendimento={att}, aviso={av}")
                rows = grupos.get((att, av), [])
                st.write("• Profissional escolhido pela Regra A/B:", [
                    g["prof_escolhido"] for g in grupos_info
                    if g["atendimento"] == att and g["aviso"] == av
                ])
                st.write("• Médicos considerados (final_pros):", sorted(set(final_pros)))
                st.write("• CÉLULAS da linha-mestre (para conferência):", rows[0].get("__cells__") if rows else None)
            else:
                st.success("✅ Nenhum par faltando: arquivo × A/B × filtro de médicos está consistente.")


                

        
        # --- (3) [OPCIONAL] Debug (temporário) — coloque após o bloco acima

             
        # --- Debug 1: Contagens e SKIPs
        total_grupos = len(grupos)
        total_AB = sum(1 for g in grupos_info if g["regra"] in ("A","B"))
        total_SKIP = sum(1 for g in grupos_info if g["regra"] == "SKIP")
        
        with st.expander("🔎 A/B vs SKIP (debug temporário)"):
            st.write("Grupos totais:", total_grupos)
            st.write("Grupos com Regra A/B:", total_AB)
            st.write("Grupos SKIP:", total_SKIP)
            if total_SKIP:
                st.table(pd.DataFrame([
                    {
                        "atendimento": g["atendimento"],
                        "aviso": g["aviso"],
                        "profissionais_parser": [ (r.get("profissional") or "") for r in g["rows"] ][:3],  # amostra
                    }
                    for g in grupos_info if g["regra"] == "SKIP"
                ]))
                 
        

        # --- Pré-visualização simples (opcional): mantém preview original por linhas ---
        registros_filtrados = registros[:] if import_all else [r for r in registros if (r.get("profissional") or "") in final_pros]
        df_preview = pd.DataFrame(registros_filtrados)
        st.subheader("Pré-visualização (DRY RUN) — nada foi gravado ainda")
        st.dataframe(df_preview, use_container_width=True, hide_index=True)

        # Info de regra consolidada
        st.markdown(
            f"<div>🔎 {len([g for g in grupos_info if g['regra'] in ('A','B')])} grupo(s) (atendimento, aviso) detectados após leitura. "
            f"Regra: {pill('1 automático por internação/aviso')}</div>",
            unsafe_allow_html=True
        )

        # ======== IMPORTAÇÃO (AGORA POR (atendimento, aviso)) ========
        colg1, colg2 = st.columns([1, 4])
        with colg1:
            if st.button("Gravar no banco", type="primary", key="import_csv_gravar"):
                total_criados = total_ignorados = total_internacoes = 0

                # Quais grupos entram segundo seleção de médicos                
                import_all = st.session_state["import_all_docs"]
                final_pros_set = set(final_pros)  # já definido acima                
                              
                st.caption(f"Pares (atendimento, aviso) considerados para gravação: {len(grupos_considerados)}")

                # 1) Atendimentos únicos
                atts_file = sorted({g["atendimento"] for g in grupos_considerados if g["atendimento"]})

                # Mapeamentos e conjuntos para busca em lote
                orig_to_norm = {att: _att_norm(att) for att in atts_file}
                norm_set = sorted({v for v in orig_to_norm.values() if v})
                num_set = sorted({_att_to_number(att) for att in atts_file if _att_to_number(att) is not None})

                # 2) Internações existentes (atendimento e numero_internacao)
                existing_map_norm_to_id = {}
                try:
                    if norm_set:
                        res_int = supabase.table("internacoes").select("id, atendimento").in_("atendimento", norm_set).execute()
                        for r in (res_int.data or []):
                            existing_map_norm_to_id[str(r["atendimento"])] = int(r["id"])
                    if num_set:
                        res_int_num = supabase.table("internacoes").select("id, numero_internacao").in_("numero_internacao", num_set).execute()
                        for r in (res_int_num.data or []):
                            k = _att_norm(str(int(float(r["numero_internacao"]))))
                            existing_map_norm_to_id[k] = int(r["id"])
                except APIError as e:
                    _sb_debug_error(e, "Falha ao buscar internações existentes.")
                    existing_map_norm_to_id = {}

                # 3) Internações faltantes (dados da LINHA MESTRE do grupo)
                to_create_int = []
                for g in grupos_considerados:
                    att = g["atendimento"]
                    na = orig_to_norm.get(att)
                    if not na:
                        continue
                    if na in existing_map_norm_to_id:
                        continue
                    m = g["master"]
                    paciente = (m.get("paciente") or "")
                    conv_total = (m.get("convenio") or "")
                    data_int = (m.get("data") or None)
                    to_create_int.append({
                        "hospital": hospital,
                        "atendimento": na,  # normalizado
                        "paciente": paciente,
                        "data_internacao": _to_ddmmyyyy(data_int) if data_int else _to_ddmmyyyy(date.today()),
                        "convenio": conv_total,
                        "numero_internacao": _att_to_number(att),
                    })

                def _chunked_insert(table_name: str, rows: list, chunk: int = 500):
                    for i in range(0, len(rows), chunk):
                        supabase.table(table_name).insert(rows[i:i+chunk]).execute()

                if to_create_int:
                    try:
                        _chunked_insert("internacoes", to_create_int, chunk=500)
                        if norm_set:
                            res_int2 = supabase.table("internacoes").select("id, atendimento").in_("atendimento", norm_set).execute()
                            for r in (res_int2.data or []):
                                existing_map_norm_to_id[str(r["atendimento"])] = int(r["id"])
                        total_internacoes = len(to_create_int)
                        invalidate_caches()
                    except APIError as e:
                        _sb_debug_error(e, "Falha ao criar internações em lote.")

                # 4) Map atendimento original -> id
                att_to_id = {att: existing_map_norm_to_id.get(orig_to_norm.get(att)) for att in atts_file}
                target_iids = sorted({iid for iid in att_to_id.values() if iid})

                # 5) Procedimentos automáticos existentes (unidade = (internacao_id, aviso))
                existing_auto = set()  # pares (iid, aviso_fmt)
                try:
                    if target_iids:
                        res_auto = (
                            supabase.table("procedimentos")
                            .select("internacao_id, aviso, is_manual")
                            .in_("internacao_id", target_iids)
                            .eq("is_manual", 0)
                            .execute()
                        )
                        for r in (res_auto.data or []):
                            iid = int(r["internacao_id"])
                            av = _fmt_id_str(r.get("aviso"))
                            if iid and av:
                                existing_auto.add((iid, av))
                except APIError as e:
                    _sb_debug_error(e, "Falha ao buscar procedimentos existentes.")

                # 6) Monta payload de novos procedimentos (1 por (internacao_id, aviso))
                to_insert_auto = []
                for g in grupos_considerados:
                    att = g["atendimento"]
                    iid = att_to_id.get(att)
                    if not iid:
                        total_ignorados += 1
                        continue

                    aviso_fmt = _fmt_id_str(g["aviso"])
                    if not aviso_fmt:
                        total_ignorados += 1
                        continue

                    if (iid, aviso_fmt) in existing_auto:
                        total_ignorados += 1
                        continue

                    m = g["master"]
                    data_norm = _to_ddmmyyyy(m.get("data"))
                    prof = g["prof_escolhido"]

                    to_insert_auto.append({
                        "internacao_id": int(iid),
                        "data_procedimento": data_norm,
                        "profissional": prof,
                        "procedimento": "Cirurgia / Procedimento",
                        "situacao": "Pendente",
                        "observacao": None,
                        "is_manual": 0,
                        "aviso": aviso_fmt,
                        "grau_participacao": None,
                    })
                    existing_auto.add((iid, aviso_fmt))  # idempotência por arquivo

                # 7) Insert em lote de procedimentos
                if to_insert_auto:
                    try:
                        _chunked_insert("procedimentos", to_insert_auto, chunk=500)
                        invalidate_caches()
                        total_criados = len(to_insert_auto)
                    except APIError as e:
                        _sb_debug_error(e, "Falha ao inserir procedimentos em lote.")

                st.success(
                    f"Concluído! Internações criadas: {total_internacoes} \n"
                    f"Automáticos criados: {total_criados} \n"
                    f"Ignorados: {total_ignorados}"
                )
                st.toast("✅ Importação concluída.", icon="✅")

        st.markdown("</div>", unsafe_allow_html=True)
        st.divider()

    # --- Seção: Cadastro manual de internação ---
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    st.subheader("➕ Cadastro manual de internação")
    cmi1, cmi2, cmi3, cmi4, cmi5 = st.columns(5)
    with cmi1: hosp_new = st.selectbox("Hospital", get_hospitais(), key="manual_new_int_hosp")
    with cmi2: att_new = st.text_input("Atendimento (único)", key="manual_new_int_att")
    with cmi3: pac_new = st.text_input("Paciente", key="manual_new_int_pac")
    with cmi4: data_new = st.date_input("Data de internação", value=date.today(), key="manual_new_int_data")
    with cmi5: conv_new = st.text_input("Convênio", key="manual_new_int_conv")
    col_btn = st.columns(6)[-1]
    with col_btn:
        if st.button("Criar internação", key="manual_btn_criar_int", type="primary"):
            if not att_new:
                st.warning("Informe o atendimento.")
            elif not get_internacao_by_atendimento(att_new).empty:
                st.error("Já existe uma internação com este atendimento (considerando zeros à esquerda).")
            else:
                nid = criar_internacao(hosp_new, att_new, pac_new, data_new.strftime("%d/%m/%Y"), conv_new)
                if nid:
                    st.toast(f"Internação criada (ID {nid}).", icon="✅")
    st.markdown("</div>", unsafe_allow_html=True)

# ============================================================
# 🔍 2) CONSULTAR
# ============================================================
with tabs[2]:
    tab_header_with_home("🔍 Consultar Internação", btn_key_suffix="consulta")
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    hlist = ["Todos"] + get_hospitais()
    filtro_hosp = st.selectbox("Filtrar hospital (consulta):", hlist)
    codigo = st.text_input("Digite o atendimento para consultar:", key="consulta_codigo", placeholder="Ex.: 0007064233 ou 7064233")
    st.markdown("</div>", unsafe_allow_html=True)

    if codigo:
        df_int = get_internacao_by_atendimento(codigo)
        if filtro_hosp != "Todos" and not df_int.empty and "hospital" in df_int.columns:
            df_int = df_int[df_int["hospital"] == filtro_hosp]
        if df_int is None or df_int.empty:
            st.warning("Nenhuma internação encontrada.")
        else:
            st.subheader("Dados da internação")
            st.dataframe(df_int, use_container_width=True, hide_index=True)
            internacao_id = int(df_int["id"].iloc[0])

            st.subheader("📝 Editar dados da internação")
            with st.container():
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    novo_paciente = st.text_input("Paciente:", value=df_int["paciente"].iloc[0] or "")
                with c2:
                    novo_convenio = st.text_input("Convênio:", value=df_int["convenio"].iloc[0] or "")
                with c3:
                    data_atual = df_int["data_internacao"].iloc[0]
                    try:
                        dt_atual = datetime.strptime(str(data_atual), "%d/%m/%Y").date()
                    except:
                        dt_atual = date.today()
                    nova_data = st.date_input("Data da internação:", value=dt_atual)
                with c4:
                    todos_hospitais = get_hospitais(include_inactive=True)
                    try:
                        idx_h = todos_hospitais.index(df_int["hospital"].iloc[0])
                    except Exception:
                        idx_h = 0 if todos_hospitais else None
                    novo_hospital = st.selectbox("Hospital:", todos_hospitais, index=idx_h if idx_h is not None else 0)
                col_save_int = st.columns(6)[-1]
                with col_save_int:
                    if st.button("💾 Salvar alterações da internação", type="primary"):
                        atualizar_internacao(
                            internacao_id,
                            paciente=novo_paciente,
                            convenio=novo_convenio,
                            data_internacao=nova_data,
                            hospital=novo_hospital
                        )
                        st.toast("Dados da internação atualizados!", icon="✅")
                        st.rerun()

            with st.expander("🗑️ Excluir esta internação"):
                st.warning("Esta ação apagará a internação e TODOS os procedimentos vinculados.")
                confirm_txt = st.text_input("Digite APAGAR para confirmar", key=f"confirm_del_int_{internacao_id}")
                col_del = st.columns(6)[-1]
                with col_del:
                    if st.button("Excluir internação", key=f"btn_del_int_{internacao_id}", type="primary"):
                        if confirm_txt.strip().upper() == "APAGAR":
                            ok = deletar_internacao(internacao_id)
                            if ok:
                                st.toast("🗑️ Internação excluída.", icon="✅")
                                st.rerun()
                            else:
                                st.stop()
                        else:
                            st.info("Confirmação inválida. Digite APAGAR.")

            try:
                res_p = supabase.table("procedimentos").select(
                    "id, data_procedimento, profissional, procedimento, situacao, observacao, aviso, grau_participacao"
                ).eq("internacao_id", internacao_id).execute()
                df_proc = pd.DataFrame(res_p.data or [])
            except APIError as e:
                _sb_debug_error(e, "Falha ao carregar procedimentos.")
                df_proc = pd.DataFrame()

            if "procedimento" not in df_proc.columns:
                df_proc["procedimento"] = "Cirurgia / Procedimento"
            for c, default in [
                ("procedimento", "Cirurgia / Procedimento"),
                ("situacao", "Pendente"),
                ("observacao", ""),
                ("aviso", ""),
                ("grau_participacao", ""),
            ]:
                if c not in df_proc.columns: df_proc[c] = default
                df_proc[c] = df_proc[c].fillna(default)

            def _safe_pt_date(s):
                try:
                    return datetime.strptime(str(s).strip(), "%d/%m/%Y").date()
                except Exception:
                    try:
                        return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()
                    except Exception:
                        return None
            df_proc["_data_dt"] = df_proc["data_procedimento"].apply(_safe_pt_date)
            df_proc = df_proc.sort_values(by=["_data_dt","id"], ascending=[True, True]).reset_index(drop=True)
            df_proc["data_procedimento"] = df_proc["_data_dt"].apply(lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else "")
            df_proc = df_proc.drop(columns=["_data_dt"])

            if "aviso" in df_proc.columns:
                df_proc["aviso"] = df_proc["aviso"].apply(_fmt_id_str)

            st.subheader("Procedimentos — Editáveis")
            edited = st.data_editor(
                df_proc,
                key="editor_proc",
                use_container_width=True, hide_index=True,
                column_config={
                    "id": st.column_config.Column("ID", disabled=True),
                    "data_procedimento": st.column_config.Column("Data", disabled=True),
                    "profissional": st.column_config.Column("Profissional", disabled=True),
                    "aviso": st.column_config.TextColumn("Aviso"),
                    "grau_participacao": st.column_config.SelectboxColumn(
                        "Grau de Participação",
                        options=[""] + GRAU_PARTICIPACAO_OPCOES,
                        required=False
                    ),
                    "procedimento": st.column_config.SelectboxColumn(
                        "Tipo de Procedimento",
                        options=PROCEDIMENTO_OPCOES,
                        required=True
                    ),
                    "situacao": st.column_config.SelectboxColumn(
                        "Situação",
                        options=STATUS_OPCOES,
                        required=True
                    ),
                    "observacao": st.column_config.TextColumn("Observações"),
                },
            )
            col_save = st.columns(6)[-1]
            with col_save:
                if st.button("💾 Salvar alterações", key="btn_save_proc", type="primary"):
                    cols_chk = ["procedimento", "situacao", "observacao", "grau_participacao", "aviso"]
                    df_compare = df_proc[["id"] + cols_chk].merge(edited[["id"] + cols_chk], on="id", suffixes=("_old", "_new"))
                    alterados = []
                    for _, row in df_compare.iterrows():
                        changed = any((str(row[c + "_old"] or "") != str(row[c + "_new"] or "")) for c in cols_chk)
                        if changed:
                            alterados.append({
                                "id": int(row["id"]),
                                "procedimento": row["procedimento_new"],
                                "situacao": row["situacao_new"],
                                "observacao": row["observacao_new"],
                                "grau_participacao": (row["grau_participacao_new"] if row["grau_participacao_new"] != "" else None),
                                "aviso": row.get("aviso_new"),
                            })
                    if not alterados:
                        st.info("Nenhuma alteração detectada.")
                    else:
                        for item in alterados:
                            atualizar_procedimento(
                                proc_id=item["id"],
                                procedimento=item["procedimento"],
                                situacao=item["situacao"],
                                observacao=item["observacao"],
                                grau_participacao=item["grau_participacao"],
                                aviso=item.get("aviso"),
                            )
                        st.toast(f"{len(alterados)} procedimento(s) atualizado(s).", icon="✅")
                        st.rerun()

            with st.expander("🗑️ Excluir cirurgia (procedimento)"):
                if df_proc.empty:
                    st.info("Não há procedimentos para excluir.")
                else:
                    for row in df_proc.itertuples(index=False):
                        proc_id = int(getattr(row, "id"))
                        data_fmt = getattr(row, "data_procedimento", "")
                        prof = getattr(row, "profissional", "") or "-"
                        tipo = getattr(row, "procedimento", "")
                        situ = getattr(row, "situacao", "")
                        c1, c2, c3, c4 = st.columns([3, 3, 3, 2])
                        with c1:
                            st.markdown(f"**ID:** {proc_id} — **Data:** {data_fmt}")
                        with c2:
                            st.markdown(f"**Profissional:** {prof}")
                        with c3:
                            st.markdown(f"**Tipo:** {tipo}<br>{pill(situ)}", unsafe_allow_html=True)
                        with c4:
                            if st.button("Excluir", key=f"del_proc_{proc_id}", help="Apagar este procedimento"):
                                ok = deletar_procedimento(proc_id)
                                if ok:
                                    st.toast(f"Procedimento {proc_id} excluído.", icon="🗑️")
                                    st.rerun()
                                else:
                                    st.stop()

            st.divider()
            st.subheader("➕ Lançar procedimento manual (permite vários no mesmo dia)")
            c1, c2, c3 = st.columns(3)
            with c1: data_proc = st.date_input("Data do procedimento", value=date.today())
            with c2:
                lista_profissionais = _listar_profissionais_cache()
                profissional = st.selectbox("Profissional", ["(selecione)"] + lista_profissionais, index=0)
            with c3:
                situacao = st.selectbox("Situação", STATUS_OPCOES, index=0)
            colp1, colp2, colp3 = st.columns(3)
            with colp1: procedimento_tipo = st.selectbox("Tipo de Procedimento", PROCEDIMENTO_OPCOES, index=0)
            with colp2: observacao = st.text_input("Observações (opcional)")
            with colp3: grau_part = st.selectbox("Grau de Participação", [""] + GRAU_PARTICIPACAO_OPCOES, index=0)
            col_add = st.columns(6)[-1]
            with col_add:
                data_internacao_str = df_int["data_internacao"].iloc[0]
                try:
                    dt_internacao = datetime.strptime(str(data_internacao_str), "%d/%m/%Y").date()
                except:
                    dt_internacao = date.today()
                if st.button("Adicionar procedimento", key="btn_add_manual", type="primary"):
                    if data_proc < dt_internacao:
                        st.error("❌ A data do procedimento não pode ser anterior à data da internação.")
                    else:
                        if profissional == "(selecione)":
                            st.error("Selecione um profissional.")
                        else:
                            criar_procedimento(
                                internacao_id, data_proc, profissional, procedimento_tipo,
                                situacao=situacao,
                                observacao=(observacao or None),
                                is_manual=1,
                                aviso=None,
                                grau_participacao=(grau_part if grau_part != "" else None),
                            )
                            st.toast("Procedimento (manual) adicionado.", icon="✅")
                            st.rerun()

            st.divider()
            st.subheader("🔎 Quitações desta internação (somente Finalizados)")
            finalizados = df_proc[df_proc["situacao"] == "Finalizado"]
            if finalizados.empty:
                st.info("Não há procedimentos finalizados nesta internação.")
            else:
                for _, r in finalizados.iterrows():
                    colA, colB, colC, colD = st.columns([2, 2, 2, 2])
                    with colA: st.markdown(f"**Data:** {r['data_procedimento']}")
                    with colB: st.markdown(f"**Profissional:** {r['profissional'] or '-'}")
                    with colC: st.markdown(f"**Aviso:** {r['aviso'] or '-'}")
                    with colD:
                        if st.button("Ver quitação", key=f"verquit_{int(r['id'])}"):
                            st.session_state["show_quit_id"] = int(r["id"])
                if "show_quit_id" in st.session_state and st.session_state["show_quit_id"]:
                    pid = int(st.session_state["show_quit_id"]); df_q = get_quitacao_by_proc_id(pid)
                    if not df_q.empty:
                        q = df_q.iloc[0]
                        aviso_fmt = _fmt_id_str(q.get("aviso"))
                        guia_amhp_fmt = _fmt_id_str(q.get("quitacao_guia_amhptiss"))
                        guia_comp_fmt = _fmt_id_str(q.get("quitacao_guia_complemento"))
                        total = float(q.get("quitacao_valor_amhptiss") or 0) + float(q.get("quitacao_valor_complemento") or 0)
                        st.markdown("---"); st.markdown("### 🧾 Detalhes da quitação")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            st.markdown(f"**Atendimento:** {q.get('atendimento','')}")
                            st.markdown(f"**Hospital:** {q.get('hospital','')}")
                            st.markdown(f"**Convênio:** {q.get('convenio') or '-'}")
                        with c2:
                            st.markdown(f"**Paciente:** {q.get('paciente','')}")
                            st.markdown(f"**Data procedimento:** {q.get('data_procedimento') or '-'}")
                            st.markdown(f"**Profissional:** {q.get('profissional') or '-'}")
                        with c3:
                            st.markdown(f"**Status:** {pill(q.get('situacao'))}", unsafe_allow_html=True)
                            st.markdown(f"**Aviso:** {aviso_fmt or '-'}")
                            st.markdown(f"**Grau participação:** {q.get('grau_participacao') or '-'}")
                        st.markdown("#### 💳 Quitação")
                        c4, c5, c6 = st.columns(3)
                        with c4:
                            st.markdown(f"**Data da quitação:** {q.get('quitacao_data') or '-'}")
                            st.markdown(f"**Guia AMHPTISS:** {guia_amhp_fmt or '-'}")
                        with c5:
                            st.markdown(f"**Valor Guia AMHPTISS:** {_format_currency_br(q.get('quitacao_valor_amhptiss'))}")
                            st.markdown(f"**Guia Complemento:** {guia_comp_fmt or '-'}")
                        with c6:
                            st.markdown(f"**Valor Guia Complemento:** {_format_currency_br(q.get('quitacao_valor_complemento'))}")
                        st.markdown(f"**Total Quitado:** **{_format_currency_br(total)}**")
                        st.markdown("**Observações da quitação:**")
                        st.write(q.get("quitacao_observacao") or "-")
                        cbot1, cbot2 = st.columns(2)
                        with cbot1:
                            if st.button("Fechar", key="fechar_quit"):
                                st.session_state["show_quit_id"] = None
                                st.rerun()
                        with cbot2:
                            if st.button("↩️ Reverter quitação", key=f"rev_{pid}", type="secondary"):
                                reverter_quitacao(pid)
                                st.toast(
                                    "Quitação revertida. Status voltou para 'Enviado para pagamento'.",
                                    icon="↩️"
                                )
                                st.session_state["show_quit_id"] = None
                                st.rerun()

# ============================================================
# 📑 3) RELATÓRIOS
# ============================================================
# ... (toda a seção de relatórios permanece idêntica ao original,
#      incluindo geração de PDFs/Excel; não foi alterada)
# Para economizar espaço, o código completo desta seção foi mantido e não modificado.

# PDF: Cirurgias por Status
if REPORTLAB_OK:
    def _pdf_cirurgias_por_status(df, filtros):
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        H1 = styles["Heading1"]; H2 = styles["Heading2"]; N = styles["BodyText"]
        from reportlab.lib.styles import ParagraphStyle
        TH = ParagraphStyle("TH", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9, leading=11, alignment=1)
        TD = ParagraphStyle("TD", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10, wordWrap="LTR")
        TD_CENTER = ParagraphStyle("TD_CENTER", parent=TD, alignment=1)
        elems = []
        elems.append(Paragraph("Relatório — Cirurgias por Status", H1)); elems.append(Spacer(1,6))
        filtros_txt = (f"Período: {filtros['ini']} a {filtros['fim']} "
                       f" Hospital: {filtros['hospital']} "
                       f" Status: {filtros['status']}")
        elems.append(Paragraph(filtros_txt, N)); elems.append(Spacer(1,8))
        total = len(df); elems.append(Paragraph(f"Total de cirurgias: <b>{total}</b>", H2))
        if total > 0 and filtros["status"] == "Todos":
            resumo = (df.groupby("situacao")["situacao"].count().sort_values(ascending=False).reset_index(name="qtd"))
            data_resumo = [["Situação", "Quantidade"]] + resumo.values.tolist()
            t_res = Table(data_resumo, hAlign="LEFT")
            t_res.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F0F0F0")),
                ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                ("ALIGN", (1,1), (-1,-1), "RIGHT"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 9),
            ]))
            elems.append(t_res); elems.append(Spacer(1,10))
        header_labels = ["Atendimento","Aviso","Convênio","Paciente","Data","Tipo","Profissional","Grau de Participação","Hospital","Situação"]
        header = [Paragraph(h, TH) for h in header_labels]
        col_widths = [2.6*cm,2.0*cm,2.8*cm,5.0*cm,2.2*cm,2.4*cm,2.8*cm,3.0*cm,2.6*cm,2.1*cm]
        def _p(v, style=TD): return Paragraph("" if v is None else str(v), style)
        data_rows = []
        for _, r in df.iterrows():
            data_rows.append([
                _p(r.get("atendimento"), TD_CENTER),
                _p(r.get("aviso"), TD_CENTER),
                _p(r.get("convenio")),
                _p(r.get("paciente")),
                _p(r.get("data_procedimento"), TD_CENTER),
                _p(r.get("procedimento")),
                _p(r.get("profissional")),
                _p(r.get("grau_participacao"), TD_CENTER),
                _p(r.get("hospital")),
                _p(r.get("situacao"), TD_CENTER),
            ])
        table = Table([header] + data_rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8EEF7")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
        ]))
        elems.append(table)
        doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        return pdf_bytes
else:
    def _pdf_cirurgias_por_status(*args, **kwargs):
        raise RuntimeError("ReportLab não está instalado no ambiente.")

# PDF: Quitações (colunas fixas)
if REPORTLAB_OK:
    def _pdf_quitacoes_colunas_fixas(df, filtros):
        need = [
            "quitacao_data","hospital","atendimento","paciente","convenio",
            "profissional","grau_participacao","data_procedimento",
            "quitacao_guia_amhptiss","quitacao_valor_amhptiss",
            "quitacao_guia_complemento","quitacao_valor_complemento",
        ]
        df = df.copy()
        for c in need:
            if c not in df.columns: df[c] = ""
        for col in ["quitacao_guia_amhptiss","quitacao_guia_complemento"]:
            df[col] = df[col].apply(_fmt_id_str)

        def _fmt_dt(s):
            d = _pt_date_to_dt(s)
            return d.strftime("%d/%m/%Y") if isinstance(d, (date, datetime)) and not pd.isna(d) else (str(s) or "")
        df["quitacao_data"] = df["quitacao_data"].apply(_fmt_dt)
        df["data_procedimento"] = df["data_procedimento"].apply(_fmt_dt)

        v_amhp = pd.to_numeric(df.get("quitacao_valor_amhptiss", 0), errors="coerce").fillna(0.0)
        v_comp = pd.to_numeric(df.get("quitacao_valor_complemento", 0), errors="coerce").fillna(0.0)
        total_amhp = float(v_amhp.sum()); total_comp = float(v_comp.sum()); total_geral = total_amhp + total_comp

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18
        )
        styles = getSampleStyleSheet()
        H1 = styles["Heading1"]; N = styles["BodyText"]
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import Table, TableStyle, Spacer, Paragraph
        TH = ParagraphStyle("TH", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8.2, leading=9.8, alignment=1)
        TD = ParagraphStyle("TD", parent=styles["Normal"], fontName="Helvetica", fontSize=7.8, leading=9.6, wordWrap="LTR")
        TD_CENTER = ParagraphStyle("TD_CENTER", parent=TD, alignment=1)
        TD_RIGHT = ParagraphStyle("TD_RIGHT", parent=TD, alignment=2)
        TD_SMALL = ParagraphStyle("TD_SMALL", parent=TD, fontSize=7.0, leading=8.6)
        def nobr(text: str) -> str:
            s = "" if text is None else str(text)
            return f"<nobr>{s}</nobr>"

        elems = []
        elems.append(Paragraph("Relatório — Quitações", H1))
        filtros_txt = f"Período da quitação: {filtros['ini']} a {filtros['fim']}  Hospital: {filtros['hospital']}"
        elems.append(Paragraph(filtros_txt, N)); elems.append(Spacer(1, 8))

        headers_raw = [
            "Quitação","Hospital", nobr("Atendimento"), "Paciente","Convênio","Profissional","Grau",
            "Data Proc.","Guia AMHPTISS","R$ AMHPTISS","Guia Compl.","R$ Compl."
        ]
        headers = [Paragraph(h, TH) for h in headers_raw]
        cols = [
            "quitacao_data","hospital","atendimento","paciente","convenio","profissional","grau_participacao",
            "data_procedimento","quitacao_guia_amhptiss","quitacao_valor_amhptiss",
            "quitacao_guia_complemento","quitacao_valor_complemento",
        ]
        col_widths = [
            2.0*cm, 2.2*cm, 2.2*cm, 4.3*cm, 2.6*cm, 3.2*cm, 1.8*cm, 2.0*cm, 2.5*cm, 2.1*cm, 2.5*cm, 2.0*cm,
        ]
        def P(v, style=TD): return Paragraph("" if v is None else str(v), style)

        data_rows = []
        for _, r in df.iterrows():
            data_rows.append([
                P(nobr(r["quitacao_data"]), TD_CENTER),
                P(r["hospital"], TD),
                P(r["atendimento"], TD_CENTER),
                P(r["paciente"], TD_SMALL),
                P(r["convenio"], TD_SMALL),
                P(r["profissional"], TD),
                P(r["grau_participacao"], TD_CENTER),
                P(nobr(r["data_procedimento"]), TD_CENTER),
                P(r["quitacao_guia_amhptiss"], TD_CENTER),
                P(_format_currency_br(r["quitacao_valor_amhptiss"]), TD_RIGHT),
                P(r["quitacao_guia_complemento"], TD_CENTER),
                P(_format_currency_br(r["quitacao_valor_complemento"]), TD_RIGHT),
            ])
        table = Table([headers] + data_rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8EEF7")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 8.2),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("ALIGN", (0,0), (-1,0), "CENTER"),
            ("ALIGN", (9,1), (9,-1), "RIGHT"),
            ("ALIGN", (11,1), (11,-1), "RIGHT"),
            ("ALIGN", (0,1), (0,-1), "CENTER"),
            ("ALIGN", (2,1), (2,-1), "CENTER"),
            ("ALIGN", (7,1), (7,-1), "CENTER"),
            ("ALIGN", (8,1), (8,-1), "CENTER"),
            ("ALIGN", (10,1), (10,-1), "CENTER"),
        ]))
        elems.append(table); elems.append(Spacer(1, 8))

        totals_data = [
            ["Total AMHPTISS:", _format_currency_br(total_amhp)],
            ["Total Complemento:", _format_currency_br(total_comp)],
            ["Total Geral:", _format_currency_br(total_geral)],
        ]
        totals_tbl = Table(totals_data, colWidths=[4.5*cm, 3.5*cm], hAlign="RIGHT")
        totals_tbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 10),
            ("ALIGN", (0,0), (0,-1), "RIGHT"),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ]))
        elems.append(totals_tbl)
        doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        return pdf_bytes

with tabs[3]:
    tab_header_with_home("📑 Relatórios — Central", btn_key_suffix="relatorios")
    # 1) Cirurgias por Status
    st.markdown("**1) Cirurgias por Status (PDF)**")
    hosp_opts = ["Todos"] + get_hospitais()
    colf1, colf2, colf3 = st.columns(3)
    with colf1:
        hosp_sel = st.selectbox("Hospital", hosp_opts, index=0, key="rel_hosp")
    with colf2:
        status_sel = st.selectbox("Status", ["Todos"] + STATUS_OPCOES, index=0, key="rel_status")
    with colf3:
        hoje = date.today()
        ini_default = hoje.replace(day=1)
        dt_ini = st.date_input("Data inicial", value=ini_default, key="rel_ini")
        dt_fim = st.date_input("Data final", value=hoje, key="rel_fim")

    df_rel = _rel_cirurgias_base_df()
    if not df_rel.empty:
        df_rel["_data_dt"] = df_rel["data_procedimento"].apply(_pt_date_to_dt)
        mask = (df_rel["_data_dt"].notna()) & (df_rel["_data_dt"] >= dt_ini) & (df_rel["_data_dt"] <= dt_fim)
        df_rel = df_rel[mask].copy()
        if hosp_sel != "Todos":
            df_rel = df_rel[df_rel["hospital"] == hosp_sel]
        if status_sel != "Todos":
            df_rel = df_rel[df_rel["situacao"] == status_sel]
        df_rel = df_rel.sort_values(by=["_data_dt","hospital","paciente","atendimento"])
        df_rel["data_procedimento"] = df_rel["_data_dt"].apply(lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else "")
        df_rel = df_rel.drop(columns=["_data_dt"])
    colc1, colc2 = st.columns(2)
    with colc1:
        if st.button("Gerar PDF (Cirurgias por Status)", key="btn_pdf_cir", type="primary"):
            if df_rel.empty:
                st.warning("Nenhum registro encontrado para os filtros informados.")
            elif not REPORTLAB_OK:
                st.error("A biblioteca 'reportlab' não está instalada no ambiente.")
            else:
                filtros = {
                    "ini": dt_ini.strftime("%d/%m/%Y"),
                    "fim": dt_fim.strftime("%d/%m/%Y"),
                    "hospital": hosp_sel,
                    "status": status_sel,
                }
                pdf_bytes = _pdf_cirurgias_por_status(df_rel, filtros)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fname = f"relatorio_cirurgias_por_status_{ts}.pdf"
                st.success(f"Relatório gerado com {len(df_rel)} registro(s).")
                st.download_button(
                    label="⬇️ Baixar PDF",
                    data=pdf_bytes,
                    file_name=fname,
                    mime="application/pdf",
                    use_container_width=True
                )
    with colc2:
        if not df_rel.empty:
            csv_bytes = df_rel.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "⬇️ Baixar CSV (fallback)",
                data=csv_bytes,
                file_name=f"cirurgias_por_status_{date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )

    st.divider()
    # 2) Quitações — PDF / CSV / Excel
    st.markdown("**2) Quitações (PDF / Excel)**")
    hosp_opts_q = ["Todos"] + get_hospitais()
    colq1, colq2 = st.columns(2)
    with colq1:
        hosp_sel_q = st.selectbox("Hospital", hosp_opts_q, index=0, key="rel_q_hosp")
    with colq2:
        hoje = date.today()
        ini_default_q = hoje.replace(day=1)
        dt_ini_q = st.date_input("Data inicial da quitação", value=ini_default_q, key="rel_q_ini")
        dt_fim_q = st.date_input("Data final da quitação", value=hoje, key="rel_q_fim")

    df_quit = _rel_quitacoes_base_df()
    if not df_quit.empty:
        df_quit["_quit_dt"] = df_quit["quitacao_data"].apply(_pt_date_to_dt)
        mask_q = (df_quit["_quit_dt"].notna()) & (df_quit["_quit_dt"] >= dt_ini_q) & (df_quit["_quit_dt"] <= dt_fim_q)
        df_quit = df_quit[mask_q].copy()
        if hosp_sel_q != "Todos":
            df_quit = df_quit[df_quit["hospital"] == hosp_sel_q]
        for col in ["quitacao_guia_amhptiss", "quitacao_guia_complemento", "aviso"]:
            if col in df_quit.columns:
                df_quit[col] = df_quit[col].apply(_fmt_id_str)
        def _fmt_dt_pt(s):
            d = _pt_date_to_dt(s)
            return d.strftime("%d/%m/%Y") if isinstance(d, (date, datetime)) and not pd.isna(d) else (str(s) or "")
        df_quit["data_procedimento"] = df_quit["data_procedimento"].apply(_fmt_dt_pt)
        df_quit["quitacao_data"] = df_quit["_quit_dt"].apply(lambda d: d.strftime("%d/%m/%Y") if pd.notna(d) else "")
        df_quit = df_quit.drop(columns=["_quit_dt"]).fillna("")
        cols_pdf = [
            "hospital","atendimento","convenio","paciente","profissional","grau_participacao",
            "data_procedimento",
            "quitacao_guia_amhptiss","quitacao_guia_complemento",
            "quitacao_valor_amhptiss","quitacao_valor_complemento",
            "quitacao_data"
        ]
        for c in cols_pdf:
            if c not in df_quit.columns:
                df_quit[c] = ""
        df_quit = df_quit.sort_values(
            by=["quitacao_data","hospital","convenio","paciente","profissional","data_procedimento"]
        ).reset_index(drop=True)

    colqb1, colqb2 = st.columns(2)
    with colqb1:
        if st.button("Gerar PDF (Quitações)", type="primary", key="btn_pdf_quit"):
            if df_quit.empty:
                st.warning("Nenhum registro de quitação encontrado para os filtros informados.")
            elif not REPORTLAB_OK:
                st.error("A biblioteca 'reportlab' não está instalada no ambiente.")
            else:
                filtros_q = {
                    "ini": dt_ini_q.strftime("%d/%m/%Y"),
                    "fim": dt_fim_q.strftime("%d/%m/%Y"),
                    "hospital": hosp_sel_q,
                }
                pdf_bytes_q = _pdf_quitacoes_colunas_fixas(df_quit, filtros_q)
                ts_q = datetime.now().strftime("%Y%m%d_%H%M%S")
                fname_q = f"relatorio_quitacoes_{ts_q}.pdf"
                st.success(f"Relatório de Quitações gerado com {len(df_quit)} registro(s).")
                st.download_button(
                    label="⬇️ Baixar PDF (Quitações)",
                    data=pdf_bytes_q,
                    file_name=fname_q,
                    mime="application/pdf",
                    use_container_width=True
                )
    with colqb2:
        if not df_quit.empty:
            csv_quit = df_quit.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "⬇️ Baixar CSV (Quitações)",
                data=csv_quit,
                file_name=f"quitacoes_{date.today().strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )
            xlsx_bytes = _excel_quitacoes_colunas_fixas(df_quit)
            st.download_button(
                "⬇️ Baixar Excel (layout do PDF)",
                data=xlsx_bytes,
                file_name=f"quitacoes_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ============================================================
# 💼 4) QUITAÇÃO (edição em lote)
# ============================================================
with tabs[4]:
    tab_header_with_home("💼 Quitação de Cirurgias", btn_key_suffix="quitacao")
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)
    hosp_opts = ["Todos"] + get_hospitais()
    hosp_sel = st.selectbox("Hospital", hosp_opts, index=0, key="quit_hosp")
    st.markdown("</div>", unsafe_allow_html=True)

    df_quit = _quitacao_pendentes_base_df()
    if hosp_sel != "Todos" and not df_quit.empty:
        df_quit = df_quit[df_quit["hospital"] == hosp_sel]

    if df_quit.empty:
        st.info("Não há cirurgias com status 'Enviado para pagamento' para quitação.")
    else:
        df_quit["quitacao_data"] = pd.to_datetime(df_quit["quitacao_data"], dayfirst=True, errors="coerce")
        for col in ["quitacao_valor_amhptiss", "quitacao_valor_complemento"]:
            df_quit[col] = pd.to_numeric(df_quit[col], errors="coerce")
        for col in ["quitacao_guia_amhptiss", "quitacao_guia_complemento"]:
            if col in df_quit.columns:
                df_quit[col] = df_quit[col].apply(_fmt_id_str)

        st.markdown("Preencha os dados e clique em **Gravar quitação(ões)**. Ao gravar, o status muda para **Finalizado**.")
        edited = st.data_editor(
            df_quit, key="editor_quit", use_container_width=True, hide_index=True,
            column_config={
                "id": st.column_config.Column("ID", disabled=True),
                "hospital": st.column_config.Column("Hospital", disabled=True),
                "atendimento": st.column_config.Column("Atendimento", disabled=True),
                "paciente": st.column_config.Column("Paciente", disabled=True),
                "convenio": st.column_config.Column("Convênio", disabled=True),
                "data_procedimento": st.column_config.Column("Data Procedimento", disabled=True),
                "profissional": st.column_config.Column("Profissional", disabled=True),
                "aviso": st.column_config.Column("Aviso", disabled=True),
                "situacao": st.column_config.Column("Situação", disabled=True),
                "quitacao_data": st.column_config.DateColumn("Data da quitação", format="DD/MM/YYYY"),
                "quitacao_guia_amhptiss": st.column_config.TextColumn("Guia AMHPTISS"),
                "quitacao_valor_amhptiss": st.column_config.NumberColumn("Valor Guia AMHPTISS", format="R$ %.2f"),
                "quitacao_guia_complemento": st.column_config.TextColumn("Guia Complemento"),
                "quitacao_valor_complemento": st.column_config.NumberColumn("Valor Guia Complemento", format="R$ %.2f"),
                "quitacao_observacao": st.column_config.TextColumn("Observações da quitação"),
            }
        )
        col_quit = st.columns(6)[-1]
        with col_quit:
            if st.button("💾 Gravar quitação(ões)", type="primary"):
                cols_chk = [
                    "quitacao_data","quitacao_guia_amhptiss","quitacao_valor_amhptiss",
                    "quitacao_guia_complemento","quitacao_valor_complemento","quitacao_observacao",
                ]
                compare = df_quit[["id"] + cols_chk].merge(edited[["id"] + cols_chk], on="id", suffixes=("_old", "_new"))
                atualizados = faltando_data = 0
                for _, row in compare.iterrows():
                    changed = any((str(row[c + "_old"] or "") != str(row[c + "_new"] or "")) for c in cols_chk)
                    if not changed: 
                        continue
                    data_q = _to_ddmmyyyy(row["quitacao_data_new"])
                    if not data_q:
                        faltando_data += 1; 
                        continue
                    guia_amhp = row["quitacao_guia_amhptiss_new"] or None
                    v_amhp = _to_float_or_none(row["quitacao_valor_amhptiss_new"])
                    guia_comp = row["quitacao_guia_complemento_new"] or None
                    v_comp = _to_float_or_none(row["quitacao_valor_complemento_new"])
                    obs_q = (row["quitacao_observacao_new"] or None)
                    quitar_procedimento(
                        proc_id=int(row["id"]),
                        data_quitacao=data_q, guia_amhptiss=guia_amhp, valor_amhptiss=v_amhp,
                        guia_complemento=guia_comp, valor_complemento=v_comp, quitacao_observacao=obs_q
                    )
                    atualizados += 1

                if faltando_data > 0 and atualizados == 0:
                    st.warning("Nenhuma quitação gravada. Preencha a **Data da quitação** para finalizar.")
                elif faltando_data > 0 and atualizados > 0:
                    st.toast(f"{atualizados} quitação(ões) gravada(s). {faltando_data} linha(s) ignoradas sem **Data da quitação**.", icon="✅")
                    st.rerun()
                else:
                    st.toast(f"{atualizados} quitação(ões) gravada(s).", icon="✅")
                    st.rerun()

# ============================================================
# ⚙️ 5) SISTEMA — Diagnósticos simples e Backups
# ============================================================
with tabs[5]:
    tab_header_with_home("⚙️ Sistema", btn_key_suffix="sistema")
    st.markdown("<div class='soft-card'>", unsafe_allow_html=True)

    # 🛡️ Backups
    st.markdown("**🛡️ Backups**")
    with st.container():
        st.caption("Gere um arquivo .zip contendo JSON e CSV de cada tabela. Opcionalmente, envie ao Supabase Storage.")
        colb1, colb2, colb3 = st.columns([2,2,2])
        with colb1:
            if st.button("🛩️ Gerar backup (ZIP)", key="btn_gen_backup", type="primary", use_container_width=True):
                with st.spinner("Gerando backup..."):
                    zip_bytes = export_tables_to_zip(["hospitals", "internacoes", "procedimentos"])
                    ts = _now_ts()
                    fname = f"backup_internacoes_{ts}.zip"
                    st.success("Backup gerado!")
                    st.download_button("⬇️ Baixar ZIP", data=zip_bytes, file_name=fname, mime="application/zip", use_container_width=True)
                    st.session_state["__last_backup_zip"] = (fname, zip_bytes)
        with colb2:
            if st.button("☁️ Enviar último backup ao Storage", key="btn_push_storage", use_container_width=True):
                last = st.session_state.get("__last_backup_zip")
                if not last:
                    st.info("Gere um backup primeiro (ou use a seção abaixo para listar/baixar do Storage).")
                else:
                    fname, zip_bytes = last
                    ok = upload_zip_to_storage(zip_bytes, fname)
                    if ok:
                        st.toast(f"Backup enviado: {fname}", icon="☁️")
        with colb3:
            st.write("")

        st.markdown("---")
        st.markdown("**☁️ Backups no Storage**")
        files = list_backups_from_storage(prefix="")
        if not files:
            st.info("Nenhum backup no Storage (ou bucket vazio).")
        else:
            for f in files[:50]:
                name = f.get("name", "")
                size = f.get("metadata", {}).get("size") or f.get("size") or 0
                updated = f.get("updated_at") or f.get("last_modified") or f.get("created_at") or "-"
                c1, c2, c3, c4 = st.columns([4, 2, 2, 2])
                with c1:
                    st.markdown(f"**{name}**")
                with c2:
                    try:
                        st.caption(f"{(int(size) or 0)/1024:.1f} KB")
                    except Exception:
                        st.caption("-")
                with c3:
                    st.caption(str(updated))
                with c4:
                    if st.button("Baixar", key=f"dl_{name}"):
                        content = download_backup_from_storage(name)
                        if content:
                            st.download_button(
                                "⬇️ Download",
                                data=content,
                                file_name=name,
                                mime="application/zip",
                                use_container_width=True,
                                key=f"dl_btn_{name}",
                            )

        st.markdown("---")
        st.markdown("**♻️ Restaurar de backup (.zip)**")
        up = st.file_uploader("Selecione o arquivo .zip do backup", type=["zip"], key="restore_zip")
        mode = st.radio("Modo de restauração", ["upsert", "replace"], index=0, help="replace apaga tudo e insere do zero (use com cautela).")
        if st.button("♻️ Restaurar", key="btn_restore", type="primary"):
            if not up:
                st.warning("Selecione um .zip primeiro.")
            else:
                with st.spinner("Restaurando..."):
                    rep = restore_from_zip(up.read(), mode=mode)
                    if rep.get("status") == "ok":
                        st.success("Restauração concluída!")
                        for d in rep.get("details", []):
                            st.write("• " + d)
                        st.toast("Caches limpos e dados restaurados.", icon="✅")
                        st.rerun()
                    else:
                        st.error("Falha na restauração.")
                        for d in rep.get("details", []):
                            st.write("• " + d)

    st.markdown("**📎 Conexão Supabase**")
    ok = True
    try:
        _ = supabase.table("hospitals").select("id", count="exact").limit(1).execute()
        st.success("Conexão OK.")
    except APIError as e:
        ok = False
        _sb_debug_error(e, "Falha ao conectar/consultar Supabase.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("**📋 Procedimentos — Lista**")
    filtro = ["Todos"] + get_hospitais()
    chosen = st.selectbox("Hospital (lista de procedimentos):", filtro, key="sys_proc_hosp")
    if st.button("Carregar procedimentos", key="btn_carregar_proc", type="primary"):
        try:
            resp = supabase.table("procedimentos").select(
                "id, internacao_id, data_procedimento, aviso, profissional, grau_participacao, procedimento, situacao, observacao"
            ).execute()
            dfp = pd.DataFrame(resp.data or [])
            if dfp.empty:
                st.info("Sem procedimentos.")
            else:
                ids = sorted(set(int(x) for x in dfp["internacao_id"].dropna().tolist()))
                resi = supabase.table("internacoes").select("id, hospital, atendimento, paciente").in_("id", ids).execute() if ids else None
                dfi = pd.DataFrame(resi.data or []) if resi else pd.DataFrame()
                df = safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left", suffixes=("", "_i"))
                if chosen != "Todos":
                    df = df[df["hospital"] == chosen]
                df = df.sort_values(by=["data_procedimento","id"], ascending=[False, False])
                st.dataframe(df, use_container_width=True, hide_index=True)
        except APIError as e:
            _sb_debug_error(e, "Falha ao carregar procedimentos.")

    st.divider()
    st.markdown("**🧾 Resumo por Profissional**")
    filtro_prof = ["Todos"] + get_hospitais()
    chosen_prof = st.selectbox("Hospital (resumo por profissional):", filtro_prof, key="sys_prof_hosp")
    try:
        resp = supabase.table("procedimentos").select("internacao_id, profissional").not_.is_("profissional", None).execute()
        dfp = pd.DataFrame(resp.data or [])
        if dfp.empty:
            st.info("Sem dados.")
        else:
            ids = sorted(set(int(x) for x in dfp["internacao_id"].dropna().tolist()))
            resi = supabase.table("internacoes").select("id, hospital").in_("id", ids).execute() if ids else None
            dfi = pd.DataFrame(resi.data or []) if resi else pd.DataFrame()
            dfm = safe_merge(dfp, dfi, left_on="internacao_id", right_on="id", how="left")
            if chosen_prof != "Todos":
                dfm = dfm[dfm["hospital"] == chosen_prof]
            df_prof = dfm.groupby("profissional")["profissional"].count().reset_index(name="total").sort_values("total", ascending=False)
            st.dataframe(df_prof, use_container_width=True, hide_index=True)
    except APIError as e:
        _sb_debug_error(e, "Falha no resumo por profissional.")

    st.divider()
    st.markdown("**💸 Resumo por Convênio**")
    filtro_conv = ["Todos"] + get_hospitais()
    chosen_conv = st.selectbox("Hospital (resumo por convênio):", filtro_conv, key="sys_conv_hosp")
    try:
        resi = supabase.table("internacoes").select("id, convenio, hospital").execute()
        dfi = pd.DataFrame(resi.data or [])
        if dfi.empty:
            st.info("Sem dados de internações.")
        else:
            if chosen_conv != "Todos":
                dfi = dfi[dfi["hospital"] == chosen_conv]
            resp = supabase.table("procedimentos").select("internacao_id").execute()
            dfp = pd.DataFrame(resp.data or [])
            if dfp.empty:
                st.info("Sem procedimentos.")
            else:
                dfp = dfp[dfp["internacao_id"].notna()]
                ids = sorted(set(int(x) for x in dfp["internacao_id"].tolist() if pd.notna(x)))
                if not ids:
                    st.info("Sem vínculos de procedimentos com internações.")
                else:
                    dfi_ids = dfi[dfi["id"].isin(ids)].copy()
                    dfm = safe_merge(dfp, dfi_ids, left_on="internacao_id", right_on="id", how="left")
                    df_conv = (
                        dfm[dfm["convenio"].notna() & (dfm["convenio"].astype(str).str.strip() != "")]
                        .groupby("convenio")["convenio"]
                        .count()
                        .reset_index(name="total")
                        .sort_values("total", ascending=False)
                    )
                    if df_conv.empty:
                        st.info("Sem dados para o resumo por convênio.")
                    else:
                        st.dataframe(df_conv, use_container_width=True, hide_index=True)
    except APIError as e:
        _sb_debug_error(e, "Falha no resumo por convênio.")

# --- Troca de aba programática ---
if st.session_state.get("goto_tab_label"):
    _switch_to_tab_by_label(st.session_state["goto_tab_label"])
    st.session_state["goto_tab_label"] = None
